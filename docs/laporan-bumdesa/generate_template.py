# -*- coding: utf-8 -*-
"""Membuat template laporan keuangan BUMDesa Lancar Jaya.

Generator ini sengaja hanya menggunakan openpyxl dan formula Excel/Google Sheets
yang umum (SUMIFS, IFERROR, VLOOKUP, MONTH, YEAR, DATE, EOMONTH, dan SUM).
Jalankan dari root repo dengan:

    python3 docs/laporan-bumdesa/generate_template.py
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "TEMPLATE_LAPORAN_KEUANGAN_BUMDESA_LANCAR_JAYA.xlsx"

# Palet utilitarian yang tetap mudah dibaca ketika diimpor ke Google Sheets.
TEAL = "06695C"
TEAL_DARK = "0A443D"
TEAL_LIGHT = "E6FBF7"
SAND = "C14B09"
SAND_LIGHT = "FFF3E0"
GREY_LIGHT = "F5F7F8"
GREY = "D2DADF"
WHITE = "FFFFFF"
GREEN = "D9F2E4"
RED = "FCE4E4"

TITLE = Font(bold=True, size=16, color=TEAL_DARK)
SUBTITLE = Font(italic=True, color="53656B")
HEADER_FONT = Font(bold=True, color=WHITE)
SECTION_FONT = Font(bold=True, color=TEAL_DARK)
TOTAL_FONT = Font(bold=True)
SMALL = Font(size=9, color="53656B")

FILL_HEADER = PatternFill("solid", fgColor=TEAL)
FILL_SECTION = PatternFill("solid", fgColor=TEAL_LIGHT)
FILL_TOTAL = PatternFill("solid", fgColor=SAND_LIGHT)
FILL_ALT = PatternFill("solid", fgColor=GREY_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN)
FILL_RED = PatternFill("solid", fgColor=RED)

THIN = Side(style="thin", color=GREY)
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RUPIAH = '#,##0;[Red]-#,##0;"-"'
DATE_FORMAT = "dd/mm/yyyy"
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

TX_DATE = "'TRANSAKSI'!$B$10:$B$1000"
TX_CODE = "'TRANSAKSI'!$E$10:$E$1000"
TX_DEBIT = "'TRANSAKSI'!$H$10:$H$1000"
TX_CREDIT = "'TRANSAKSI'!$I$10:$I$1000"
TX_MONTH = "'TRANSAKSI'!$J$10:$J$1000"
TX_YEAR = "'TRANSAKSI'!$K$10:$K$1000"


# Kode akun detail yang sudah dibersihkan dari trailing space dan duplikat.
# Kolom ketiga adalah klasifikasi laporan, kolom keempat saldo normal.
COA = [
    ("1.1.01.01", "Kas Tunai", "Aset", "Debit"),
    ("1.1.01.02", "Kas di Bank BSI", "Aset", "Debit"),
    ("1.1.01.03", "Kas di Bank Mandiri", "Aset", "Debit"),
    ("1.1.01.04", "Kas di Bank BRI", "Aset", "Debit"),
    ("1.1.01.05", "Kas di Bank BPD", "Aset", "Debit"),
    ("1.1.01.06", "Kas di Bank Jatim", "Aset", "Debit"),
    ("1.1.01.98", "Kas Kecil (Petty Cash)", "Aset", "Debit"),
    ("1.1.03.01", "Piutang Usaha", "Aset", "Debit"),
    ("1.1.05.01", "Persediaan Barang Dagangan", "Aset", "Debit"),
    ("1.1.05.02", "Persediaan Bahan Baku", "Aset", "Debit"),
    ("1.1.07.01", "Sewa Dibayar Dimuka", "Aset", "Debit"),
    ("1.3.01.01", "Tanah", "Aset", "Debit"),
    ("1.3.02.01", "Kendaraan", "Aset", "Debit"),
    ("1.3.03.01", "Peralatan dan Mesin", "Aset", "Debit"),
    ("1.3.04.01", "Meubelair", "Aset", "Debit"),
    ("1.3.05.01", "Gedung dan Bangunan", "Aset", "Debit"),
    ("1.3.07.01", "Akum. Penyusutan Kendaraan", "Aset", "Kredit"),
    ("1.3.07.02", "Akum. Penyusutan Peralatan dan Mesin", "Aset", "Kredit"),
    ("1.3.07.03", "Akum. Penyusutan Meubelair", "Aset", "Kredit"),
    ("1.3.07.04", "Akum. Penyusutan Gedung dan Bangunan", "Aset", "Kredit"),
    ("1.3.99.01", "Aset Biologis - Sapi", "Aset", "Debit"),
    ("1.3.99.04", "Aset Biologis - Ayam", "Aset", "Debit"),
    ("1.3.99.05", "Aset Biologis - Padi", "Aset", "Debit"),
    ("1.3.99.06", "Aset Biologis - Jagung", "Aset", "Debit"),
    ("1.4.01.01", "Software", "Aset", "Debit"),
    ("2.1.01.01", "Utang Usaha", "Kewajiban", "Kredit"),
    ("2.1.03.01", "Utang Gaji dan Upah", "Kewajiban", "Kredit"),
    ("2.1.04.01", "Utang Listrik", "Kewajiban", "Kredit"),
    ("2.1.05.01", "Utang Pihak Ketiga Jangka Pendek", "Kewajiban", "Kredit"),
    ("2.2.01.01", "Utang Bank", "Kewajiban", "Kredit"),
    ("3.1.01.01", "Penyertaan Modal Desa", "Ekuitas", "Kredit"),
    ("3.1.02.01", "Penyertaan Modal Masyarakat A", "Ekuitas", "Kredit"),
    ("3.1.02.02", "Penyertaan Modal Masyarakat B", "Ekuitas", "Kredit"),
    ("3.3.01.04", "Saldo Awal", "Ekuitas", "Kredit"),
    ("3.4.01.01", "Modal Donasi dan Sumbangan", "Ekuitas", "Kredit"),
    ("4.1.02.01", "Pendapatan Pengelolaan Air Bersih", "Pendapatan", "Kredit"),
    ("4.1.03.01", "Pendapatan Pengelolaan Sampah", "Pendapatan", "Kredit"),
    ("4.1.04.02", "Pendapatan Sewa Toko dan Kios", "Pendapatan", "Kredit"),
    ("4.1.04.03", "Pendapatan Sewa Gedung", "Pendapatan", "Kredit"),
    ("4.1.04.06", "Pendapatan Sewa Jaring", "Pendapatan", "Kredit"),
    ("4.1.05.01", "Pendapatan Jasa Bayar Listrik", "Pendapatan", "Kredit"),
    ("4.1.05.03", "Pendapatan Jasa Bayar BPJS", "Pendapatan", "Kredit"),
    ("4.1.08.01", "Pendapatan Simpan Pinjam", "Pendapatan", "Kredit"),
    ("4.1.11.01", "Pendapatan Komisi", "Pendapatan", "Kredit"),
    ("4.1.12.01", "Pendapatan Penjualan Jagung", "Pendapatan", "Kredit"),
    ("4.1.12.02", "Pendapatan Penjualan Padi", "Pendapatan", "Kredit"),
    ("4.1.13.05", "Pendapatan Penjualan Telur Ayam", "Pendapatan", "Kredit"),
    ("4.1.14.01", "Pendapatan Penjualan Lele", "Pendapatan", "Kredit"),
    ("6.1.02.01", "Beban Alat Tulis Kantor (ATK)", "Beban", "Debit"),
    ("6.1.02.02", "Beban Foto Copy", "Beban", "Debit"),
    ("6.1.02.03", "Beban Konsumsi Rapat", "Beban", "Debit"),
    ("6.1.04.01", "Beban Listrik", "Beban", "Debit"),
    ("6.1.04.02", "Beban Telepon/Internet", "Beban", "Debit"),
    ("6.1.07.02", "Beban Penyusutan Kendaraan", "Beban", "Debit"),
    ("6.1.07.03", "Beban Penyusutan Peralatan dan Mesin", "Beban", "Debit"),
    ("6.1.07.04", "Beban Penyusutan Meubelair", "Beban", "Debit"),
    ("6.1.07.05", "Beban Penyusutan Gedung dan Bangunan", "Beban", "Debit"),
    ("6.1.99.03", "Beban Perjalanan Dinas", "Beban", "Debit"),
    ("6.1.99.05", "Beban Jamuan Tamu", "Beban", "Debit"),
    ("6.2.01.03", "Beban Gaji Penasihat", "Beban", "Debit"),
    ("6.2.01.04", "Beban Gaji Pengawas", "Beban", "Debit"),
    ("6.2.01.05", "Beban Gaji Direktur", "Beban", "Debit"),
    ("6.2.01.07", "Beban Gaji Bendahara", "Beban", "Debit"),
    ("6.2.02.02", "Beban Perbaikan dan Renovasi", "Beban", "Debit"),
    ("6.2.99.02", "Beban Sewa Lokasi", "Beban", "Debit"),
    ("6.2.99.04", "Beban Pakan Ternak", "Beban", "Debit"),
    ("6.2.99.05", "Beban Pupuk dan Obat-obatan", "Beban", "Debit"),
    ("6.2.99.06", "Beban Peralatan dan Mesin", "Beban", "Debit"),
    ("6.2.99.07", "Beban Vaksinasi dan Obat", "Beban", "Debit"),
    ("6.2.99.99", "Beban Operasional Lainnya", "Beban", "Debit"),
    ("7.2.01.01", "Beban Administrasi Bank", "Beban", "Debit"),
    ("7.3.01.03", "Beban Pajak Daerah", "Beban", "Debit"),
]

ACCOUNT_BY_CODE = {row[0]: row[1] for row in COA}
INCOME = [row for row in COA if row[2] == "Pendapatan"]
EXPENSE = [row for row in COA if row[2] == "Beban"]
ASSETS = [row for row in COA if row[2] == "Aset"]
LIABILITIES = [row for row in COA if row[2] == "Kewajiban"]
EQUITY = [row for row in COA if row[2] == "Ekuitas"]
CASH_CODES = [row for row in ASSETS if row[0].startswith("1.1.01")]


def add_defined_name(wb, name, attr_text):
    """Tambahkan atau ganti defined name secara kompatibel dengan openpyxl 3.1."""
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name=name, attr_text=attr_text))


def style_header(ws, row, start_col, labels):
    for offset, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + offset, value=label)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = GRID


def style_row(ws, row, start_col, end_col, fill=None, font=None):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = GRID
        if fill:
            cell.fill = fill
        if font:
            cell.font = font


def section_row(ws, row, start_col, end_col, text):
    ws.cell(row=row, column=start_col, value=text).font = SECTION_FONT
    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).fill = FILL_SECTION
        ws.cell(row=row, column=col).border = GRID


def money(cell):
    cell.number_format = RUPIAH
    cell.alignment = RIGHT


def add_table(ws, ref, name):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def signed_period_formula(code_ref, positive_column, negative_column, start_expr, end_expr):
    """Saldo laporan untuk periode: kredit-debit atau debit-kredit."""
    return (
        f'=SUMIFS({positive_column},{TX_CODE},{code_ref},{TX_DATE},">="&{start_expr},'
        f'{TX_DATE},"<="&{end_expr})-SUMIFS({negative_column},{TX_CODE},{code_ref},'
        f'{TX_DATE},">="&{start_expr},{TX_DATE},"<="&{end_expr})'
    )


def signed_to_date_formula(code_ref, positive_column, negative_column, end_expr):
    return (
        f'=SUMIFS({positive_column},{TX_CODE},{code_ref},{TX_DATE},"<="&{end_expr})-'
        f'SUMIFS({negative_column},{TX_CODE},{code_ref},{TX_DATE},"<="&{end_expr})'
    )


def monthly_account_formula(code, category, month_number):
    if category == "Pendapatan":
        positive, negative = TX_CREDIT, TX_DEBIT
    else:
        positive, negative = TX_DEBIT, TX_CREDIT
    return (
        f'=SUMIFS({positive},{TX_CODE},"{code}",{TX_MONTH},{month_number},{TX_YEAR},TahunSel)-'
        f'SUMIFS({negative},{TX_CODE},"{code}",{TX_MONTH},{month_number},{TX_YEAR},TahunSel)'
    )


def build_pengaturan(ws):
    ws.sheet_properties.tabColor = TEAL_DARK
    widths = {"A": 3, "B": 32, "C": 30, "D": 35, "E": 25, "F": 18, "G": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "BUMDESA LANCAR JAYA"
    ws["B2"].font = Font(bold=True, size=18, color=TEAL_DARK)
    ws["B3"] = "Sistem laporan keuangan multi-bulan — Desa Curahdringu, Kecamatan Tongas, Kabupaten Probolinggo"
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:E2")
    ws.merge_cells("B3:G3")

    identity = [
        ("Badan Usaha", "LANCAR JAYA"),
        ("Desa", "CURAHDRINGU"),
        ("Kecamatan", "TONGAS"),
        ("Kabupaten", "PROBOLINGGO"),
        ("Direktur", "DITA TIA MUKARROMAH"),
        ("Bendahara", "ROBIATUL HUSNA"),
        ("Penasihat", "................................"),
        ("Pengawas", "................................"),
    ]
    for row, (label, value) in enumerate(identity, start=5):
        ws.cell(row=row, column=2, value=label).font = Font(bold=True)
        ws.cell(row=row, column=3, value=value)
        style_row(ws, row, 2, 3)

    ws["B15"] = "KONTROL PERIODE (ubah di sini)"
    ws["B15"].font = Font(bold=True, color=SAND)
    ws["B16"] = "Bulan aktif (1-12)"
    ws["C16"] = 9
    ws["B17"] = "Tahun aktif"
    ws["C17"] = 2025
    ws["B18"] = "Nama bulan"
    ws["C18"] = '=TEXT(DATE(TahunSel,BulanSel,1),"mmmm")'
    ws["B19"] = "Periode aktif"
    ws["C19"] = "=DATE(TahunSel,BulanSel,1)"
    ws["C19"].number_format = "mmmm yyyy"
    for row in range(16, 20):
        style_row(ws, row, 2, 3, fill=FILL_TOTAL if row in (16, 17) else None)
        ws.cell(row=row, column=2).font = Font(bold=True)
    ws["C16"].font = TOTAL_FONT
    ws["C17"].font = TOTAL_FONT
    ws["C16"].alignment = CENTER
    ws["C17"].alignment = CENTER

    month_dv = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8,9,10,11,12"', allow_blank=False)
    year_dv = DataValidation(type="list", formula1='"2024,2025,2026,2027,2028,2029,2030,2031,2032,2033,2034,2035"', allow_blank=False)
    ws.add_data_validation(month_dv)
    ws.add_data_validation(year_dv)
    month_dv.add(ws["C16"])
    year_dv.add(ws["C17"])

    ws["B21"] = "DAFTAR UNIT USAHA (boleh diedit)"
    ws["B21"].font = Font(bold=True, color=SAND)
    units = [
        "Pusat", "Unit Pertanian", "Unit Peternakan", "Unit Air Bersih",
        "Unit Simpan Pinjam", "Unit Sewa Aset", "Unit Pelayanan Jasa",
    ]
    for row, unit in enumerate(units, start=22):
        ws.cell(row=row, column=2, value=unit)
        style_row(ws, row, 2, 2)
    add_defined_name(ws.parent, "BulanSel", "PENGATURAN!$C$16")
    add_defined_name(ws.parent, "TahunSel", "PENGATURAN!$C$17")
    add_defined_name(ws.parent, "UnitUsaha", "PENGATURAN!$B$22:$B$28")


def build_akun(ws, wb):
    ws.sheet_properties.tabColor = TEAL
    for col, width in {"A": 16, "B": 48, "C": 16, "D": 15}.items():
        ws.column_dimensions[col].width = width
    ws["A1"] = "KODE AKUN"
    ws["B1"] = "NAMA AKUN"
    ws["C1"] = "KLASIFIKASI"
    ws["D1"] = "SALDO NORMAL"
    style_header(ws, 1, 1, ["KODE AKUN", "NAMA AKUN", "KLASIFIKASI", "SALDO NORMAL"])
    for row, (code, name, category, normal) in enumerate(COA, start=2):
        ws.cell(row=row, column=1, value=code).number_format = "@"
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=category)
        ws.cell(row=row, column=4, value=normal)
        style_row(ws, row, 1, 4, fill=FILL_ALT if row % 2 == 0 else None)
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=2).alignment = LEFT
    add_table(ws, f"A1:D{len(COA) + 1}", "DaftarAkun")
    # Rentang sampai baris 200 memudahkan pemilihan kode yang sudah dimasukkan.
    add_defined_name(wb, "KodeAkun", "AKUN!$A$2:$A$200")
    ws["B78"] = "Gunakan akun yang tersedia. Untuk akun baru permanen, tambahkan tuple ke COA di generate_template.py lalu jalankan ulang agar semua laporan mendapat barisnya."
    ws["B78"].font = SMALL
    ws.merge_cells("B78:D78")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(COA) + 1}"


def build_transaksi(ws, wb):
    ws.sheet_properties.tabColor = SAND
    widths = {"A": 8, "B": 14, "C": 22, "D": 18, "E": 16, "F": 44, "G": 58, "H": 17, "I": 17, "J": 10, "K": 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws["A2"] = "TRANSAKSI JURNAL TAHUNAN"
    ws["A2"].font = TITLE
    ws["A3"] = "Satu baris adalah satu akun. Untuk satu transaksi, isi minimal satu baris debet dan satu baris kredit dengan tanggal, unit, serta uraian yang sama."
    ws["A3"].font = SUBTITLE
    ws.merge_cells("A2:K2")
    ws.merge_cells("A3:K3")
    headers = ["No", "Tanggal", "Unit Usaha", "Jenis", "Kode Akun", "Nama Akun", "Uraian", "Debet", "Kredit", "Bulan", "Tahun"]
    style_header(ws, 9, 1, headers)

    # Formula siap pakai pada 1000 baris. Kolom input sengaja tetap kosong pada baris 10:1000.
    for row in range(10, 1001):
        ws.cell(row=row, column=1, value=f'=IF($B{row}="","",ROW()-9)')
        ws.cell(row=row, column=6, value=f'=IF($E{row}="","",IFERROR(VLOOKUP($E{row},AKUN!$A$2:$B$200,2,FALSE),"Kode tidak ditemukan"))')
        ws.cell(row=row, column=10, value=f'=IF($B{row}="","",MONTH($B{row}))')
        ws.cell(row=row, column=11, value=f'=IF($B{row}="","",YEAR($B{row}))')
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = GRID
            if row % 2 == 0:
                cell.fill = FILL_ALT
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2).number_format = DATE_FORMAT
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=8).number_format = RUPIAH
        ws.cell(row=row, column=9).number_format = RUPIAH
        ws.cell(row=row, column=8).alignment = RIGHT
        ws.cell(row=row, column=9).alignment = RIGHT
        ws.cell(row=row, column=10).alignment = CENTER
        ws.cell(row=row, column=11).alignment = CENTER

    sample = [
        (datetime(2025, 9, 1), "Unit Pertanian", "Operasional", "1.1.01.01", "[CONTOH] Saldo awal kas usaha pertanian", 97_950_000, None),
        (datetime(2025, 9, 1), "Unit Pertanian", "Operasional", "3.3.01.04", "[CONTOH] Saldo awal kas usaha pertanian", None, 97_950_000),
        (datetime(2025, 9, 5), "Unit Pertanian", "Operasional", "1.1.01.01", "[CONTOH] Penjualan padi", 4_000_000, None),
        (datetime(2025, 9, 5), "Unit Pertanian", "Operasional", "4.1.12.02", "[CONTOH] Penjualan padi", None, 4_000_000),
        (datetime(2025, 9, 10), "Unit Pertanian", "Operasional", "1.1.01.01", "[CONTOH] Penjualan jagung", 3_000_000, None),
        (datetime(2025, 9, 10), "Unit Pertanian", "Operasional", "4.1.12.01", "[CONTOH] Penjualan jagung", None, 3_000_000),
        (datetime(2025, 9, 19), "Unit Pertanian", "Operasional", "6.2.99.99", "[CONTOH] Upah pemupukan 2 orang", 200_000, None),
        (datetime(2025, 9, 19), "Unit Pertanian", "Operasional", "1.1.01.01", "[CONTOH] Upah pemupukan 2 orang", None, 200_000),
        (datetime(2025, 9, 19), "Unit Pertanian", "Operasional", "6.2.99.05", "[CONTOH] Pembelian pupuk phonska 1 sak", 75_000, None),
        (datetime(2025, 9, 19), "Unit Pertanian", "Operasional", "1.1.01.01", "[CONTOH] Pembelian pupuk phonska 1 sak", None, 75_000),
        (datetime(2025, 9, 22), "Unit Pertanian", "Operasional", "6.1.04.01", "[CONTOH] Pembayaran listrik", 150_000, None),
        (datetime(2025, 9, 22), "Unit Pertanian", "Operasional", "1.1.01.01", "[CONTOH] Pembayaran listrik", None, 150_000),
        (datetime(2025, 9, 30), "Unit Pertanian", "Penyesuaian", "6.2.99.02", "[CONTOH] Penyesuaian beban sewa bulan berjalan", 500_000, None),
        (datetime(2025, 9, 30), "Unit Pertanian", "Penyesuaian", "1.1.07.01", "[CONTOH] Penyesuaian beban sewa bulan berjalan", None, 500_000),
        (datetime(2025, 9, 30), "Unit Pertanian", "Penyesuaian", "6.1.07.03", "[CONTOH] Penyusutan peralatan dan mesin", 300_000, None),
        (datetime(2025, 9, 30), "Unit Pertanian", "Penyesuaian", "1.3.07.02", "[CONTOH] Penyusutan peralatan dan mesin", None, 300_000),
    ]
    for row, record in enumerate(sample, start=10):
        date, unit, kind, code, description, debit, credit = record
        ws.cell(row=row, column=2, value=date)
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=kind)
        ws.cell(row=row, column=5, value=code).number_format = "@"
        ws.cell(row=row, column=7, value=description)
        ws.cell(row=row, column=8, value=debit)
        ws.cell(row=row, column=9, value=credit)
        ws.cell(row=row, column=2).number_format = DATE_FORMAT
        if debit is not None:
            ws.cell(row=row, column=8).number_format = RUPIAH
        if credit is not None:
            ws.cell(row=row, column=9).number_format = RUPIAH

    add_table(ws, "A9:K1000", "TabelTransaksi")
    ws.auto_filter.ref = "A9:K1000"
    ws.freeze_panes = "A10"

    unit_dv = DataValidation(type="list", formula1="=UnitUsaha", allow_blank=True)
    kind_dv = DataValidation(type="list", formula1='"Operasional,Penyesuaian"', allow_blank=True)
    code_dv = DataValidation(type="list", formula1="=KodeAkun", allow_blank=True)
    date_dv = DataValidation(type="date", operator="between", formula1="DATE(2020,1,1)", formula2="DATE(2040,12,31)", allow_blank=True)
    for validation in (unit_dv, kind_dv, code_dv, date_dv):
        ws.add_data_validation(validation)
    unit_dv.add("C10:C1000")
    kind_dv.add("D10:D1000")
    code_dv.add("E10:E1000")
    date_dv.add("B10:B1000")


def build_bulanan(ws):
    ws.sheet_properties.tabColor = "2F75B5"
    for col, width in {"A": 16, "B": 43, "C": 15, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14, "O": 14, "P": 16}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "RINGKASAN BULANAN PER AKUN"
    ws["B2"].font = TITLE
    ws["B3"] = '="Tahun aktif: "&TahunSel'
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:P2")
    style_header(ws, 6, 1, ["KODE AKUN", "NAMA AKUN", "JENIS", "Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des", "TOTAL TAHUN"])
    row = 7
    section_row(ws, row, 1, 16, "PENDAPATAN")
    row += 1
    income_start = row
    for code, name, category, _ in INCOME:
        ws.cell(row=row, column=1, value=code).number_format = "@"
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=category)
        for month, col in enumerate(range(4, 16), start=1):
            ws.cell(row=row, column=col, value=monthly_account_formula(code, category, month))
            money(ws.cell(row=row, column=col))
        ws.cell(row=row, column=16, value=f"=SUM(D{row}:O{row})")
        money(ws.cell(row=row, column=16))
        style_row(ws, row, 1, 16, fill=FILL_ALT if row % 2 == 0 else None)
        row += 1
    income_total = row
    ws.cell(row=row, column=2, value="Total Pendapatan")
    for col in range(4, 17):
        letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=SUM({letter}{income_start}:{letter}{row - 1})")
        money(ws.cell(row=row, column=col))
    style_row(ws, row, 1, 16, fill=FILL_TOTAL, font=TOTAL_FONT)
    row += 2
    section_row(ws, row, 1, 16, "BEBAN")
    row += 1
    expense_start = row
    for code, name, category, _ in EXPENSE:
        ws.cell(row=row, column=1, value=code).number_format = "@"
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=category)
        for month, col in enumerate(range(4, 16), start=1):
            ws.cell(row=row, column=col, value=monthly_account_formula(code, category, month))
            money(ws.cell(row=row, column=col))
        ws.cell(row=row, column=16, value=f"=SUM(D{row}:O{row})")
        money(ws.cell(row=row, column=16))
        style_row(ws, row, 1, 16, fill=FILL_ALT if row % 2 == 0 else None)
        row += 1
    expense_total = row
    ws.cell(row=row, column=2, value="Total Beban")
    for col in range(4, 17):
        letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=SUM({letter}{expense_start}:{letter}{row - 1})")
        money(ws.cell(row=row, column=col))
    style_row(ws, row, 1, 16, fill=FILL_TOTAL, font=TOTAL_FONT)
    row += 2
    profit_row = row
    ws.cell(row=row, column=2, value="SURPLUS / (RUGI) PER BULAN")
    for col in range(4, 17):
        letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"={letter}{income_total}-{letter}{expense_total}")
        money(ws.cell(row=row, column=col))
    style_row(ws, row, 1, 16, fill=FILL_TOTAL, font=TOTAL_FONT)
    row += 2
    cash_row = row
    ws.cell(row=row, column=2, value="SALDO KAS AKHIR (akun 1.1.01*)")
    for month, col in enumerate(range(4, 16), start=1):
        end_expr = f"EOMONTH(DATE(TahunSel,{month},1),0)"
        formula = f'=SUMIFS({TX_DEBIT},{TX_CODE},"1.1.01*",{TX_DATE},"<="&{end_expr})-SUMIFS({TX_CREDIT},{TX_CODE},"1.1.01*",{TX_DATE},"<="&{end_expr})'
        ws.cell(row=row, column=col, value=formula)
        money(ws.cell(row=row, column=col))
    ws.cell(row=row, column=16, value="=O" + str(row))
    money(ws.cell(row=row, column=16))
    style_row(ws, row, 1, 16, fill=FILL_SECTION, font=TOTAL_FONT)
    ws.cell(row=row + 2, column=2, value="Semua angka bersumber dari TRANSAKSI berdasarkan kode akun, bulan, dan tahun; bukan posisi baris.").font = SMALL
    ws.merge_cells(start_row=row + 2, start_column=2, end_row=row + 2, end_column=16)
    ws.freeze_panes = "D7"
    return {"income_total": income_total, "expense_total": expense_total, "profit": profit_row, "cash": cash_row}


def build_laba_rugi(ws):
    ws.sheet_properties.tabColor = "70AD47"
    for col, width in {"A": 16, "B": 48, "C": 20, "D": 20}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "LAPORAN LABA RUGI"
    ws["B2"].font = TITLE
    ws["B3"] = '="Periode: "&TEXT(DATE(TahunSel,BulanSel,1),"mmmm yyyy")'
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:D2")
    style_header(ws, 6, 1, ["KODE AKUN", "URAIAN", "BULAN INI", "S/D BULAN INI"])
    row = 7
    section_row(ws, row, 1, 4, "PENDAPATAN")
    row += 1
    income_start = row
    for code, name, category, _ in INCOME:
        ws.cell(row=row, column=1, value=code).number_format = "@"
        ws.cell(row=row, column=2, value=name)
        current_end = "EOMONTH(DATE(TahunSel,BulanSel,1),0)"
        current_start = "DATE(TahunSel,BulanSel,1)"
        ytd_start = "DATE(TahunSel,1,1)"
        ws.cell(row=row, column=3, value=signed_period_formula(f"$A{row}", TX_CREDIT, TX_DEBIT, current_start, current_end))
        ws.cell(row=row, column=4, value=signed_period_formula(f"$A{row}", TX_CREDIT, TX_DEBIT, ytd_start, current_end))
        money(ws.cell(row=row, column=3))
        money(ws.cell(row=row, column=4))
        style_row(ws, row, 1, 4, fill=FILL_ALT if row % 2 == 0 else None)
        row += 1
    income_total = row
    ws.cell(row=row, column=2, value="Total Pendapatan")
    ws.cell(row=row, column=3, value=f"=SUM(C{income_start}:C{row - 1})")
    ws.cell(row=row, column=4, value=f"=SUM(D{income_start}:D{row - 1})")
    money(ws.cell(row=row, column=3))
    money(ws.cell(row=row, column=4))
    style_row(ws, row, 1, 4, fill=FILL_TOTAL, font=TOTAL_FONT)
    row += 2
    section_row(ws, row, 1, 4, "BEBAN")
    row += 1
    expense_start = row
    for code, name, category, _ in EXPENSE:
        ws.cell(row=row, column=1, value=code).number_format = "@"
        ws.cell(row=row, column=2, value=name)
        current_end = "EOMONTH(DATE(TahunSel,BulanSel,1),0)"
        current_start = "DATE(TahunSel,BulanSel,1)"
        ytd_start = "DATE(TahunSel,1,1)"
        ws.cell(row=row, column=3, value=signed_period_formula(f"$A{row}", TX_DEBIT, TX_CREDIT, current_start, current_end))
        ws.cell(row=row, column=4, value=signed_period_formula(f"$A{row}", TX_DEBIT, TX_CREDIT, ytd_start, current_end))
        money(ws.cell(row=row, column=3))
        money(ws.cell(row=row, column=4))
        style_row(ws, row, 1, 4, fill=FILL_ALT if row % 2 == 0 else None)
        row += 1
    expense_total = row
    ws.cell(row=row, column=2, value="Total Beban")
    ws.cell(row=row, column=3, value=f"=SUM(C{expense_start}:C{row - 1})")
    ws.cell(row=row, column=4, value=f"=SUM(D{expense_start}:D{row - 1})")
    money(ws.cell(row=row, column=3))
    money(ws.cell(row=row, column=4))
    style_row(ws, row, 1, 4, fill=FILL_TOTAL, font=TOTAL_FONT)
    row += 2
    profit_row = row
    ws.cell(row=row, column=2, value="SURPLUS / (RUGI)")
    ws.cell(row=row, column=3, value=f"=C{income_total}-C{expense_total}")
    ws.cell(row=row, column=4, value=f"=D{income_total}-D{expense_total}")
    money(ws.cell(row=row, column=3))
    money(ws.cell(row=row, column=4))
    style_row(ws, row, 1, 4, fill=FILL_TOTAL, font=TOTAL_FONT)
    sign_row = row + 4
    ws.cell(row=sign_row, column=2, value='="Curahdringu, "&TEXT(EOMONTH(DATE(TahunSel,BulanSel,1),0),"dd mmmm yyyy")')
    ws.cell(row=sign_row + 2, column=2, value='=PENGATURAN!$C$9')
    ws.cell(row=sign_row + 2, column=2).font = Font(bold=True)
    ws.cell(row=sign_row + 3, column=2, value="Direktur")
    ws.cell(row=sign_row + 2, column=4, value='=PENGATURAN!$C$10')
    ws.cell(row=sign_row + 2, column=4).font = Font(bold=True)
    ws.cell(row=sign_row + 3, column=4, value="Bendahara")
    ws.cell(row=sign_row + 5, column=2, value="Penasihat/Pengawas dapat diisi di PENGATURAN.").font = SMALL
    ws.freeze_panes = "A7"
    return {"income_total": income_total, "expense_total": expense_total, "profit": profit_row}


def build_arus_kas(ws):
    ws.sheet_properties.tabColor = "00A6A6"
    for col, width in {"A": 4, "B": 16, "C": 42, "D": 20, "E": 20, "F": 20, "G": 20, "H": 20}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "LAPORAN ARUS KAS"
    ws["B2"].font = TITLE
    ws["B3"] = '="Periode: "&TEXT(DATE(TahunSel,BulanSel,1),"mmmm yyyy")'
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:H2")
    style_header(ws, 6, 2, ["KODE AKUN KAS", "NAMA REKENING KAS", "SALDO AWAL BULAN", "PENERIMAAN KAS", "PENGELUARAN KAS", "ARUS BERSIH", "SALDO AKHIR"])
    row = 7
    section_row(ws, row, 2, 8, "REKENING KAS (kriteria kode akun 1.1.01*)")
    row += 1
    cash_start = row
    start_expr = "DATE(TahunSel,BulanSel,1)"
    end_expr = "EOMONTH(DATE(TahunSel,BulanSel,1),0)"
    for code, name, _, _ in CASH_CODES:
        ws.cell(row=row, column=2, value=code).number_format = "@"
        ws.cell(row=row, column=3, value=name)
        ws.cell(row=row, column=4, value=f'=SUMIFS({TX_DEBIT},{TX_CODE},$B{row},{TX_DATE},"<"&{start_expr})-SUMIFS({TX_CREDIT},{TX_CODE},$B{row},{TX_DATE},"<"&{start_expr})')
        ws.cell(row=row, column=5, value=f'=SUMIFS({TX_DEBIT},{TX_CODE},$B{row},{TX_DATE},">="&{start_expr},{TX_DATE},"<="&{end_expr})')
        ws.cell(row=row, column=6, value=f'=SUMIFS({TX_CREDIT},{TX_CODE},$B{row},{TX_DATE},">="&{start_expr},{TX_DATE},"<="&{end_expr})')
        ws.cell(row=row, column=7, value=f"=E{row}-F{row}")
        ws.cell(row=row, column=8, value=f"=D{row}+G{row}")
        for col in range(4, 9):
            money(ws.cell(row=row, column=col))
        style_row(ws, row, 2, 8, fill=FILL_ALT if row % 2 == 0 else None)
        row += 1
    cash_total = row
    ws.cell(row=row, column=3, value="TOTAL KAS")
    for col in range(4, 9):
        letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=SUM({letter}{cash_start}:{letter}{row - 1})")
        money(ws.cell(row=row, column=col))
    style_row(ws, row, 2, 8, fill=FILL_TOTAL, font=TOTAL_FONT)
    ws.cell(row=row + 2, column=2, value="Penerimaan/pengeluaran dihitung dari Debet/Kredit rekening kas, bukan dari teks uraian transaksi.").font = SMALL
    ws.merge_cells(start_row=row + 2, start_column=2, end_row=row + 2, end_column=8)
    ws.freeze_panes = "D7"
    return {"total": cash_total}


def build_neraca_saldo(ws, laba_rugi_rows):
    ws.sheet_properties.tabColor = "8064A2"
    for col, width in {"A": 16, "B": 48, "C": 20, "D": 20, "E": 20, "F": 20}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "NERACA SALDO"
    ws["B2"].font = TITLE
    ws["B3"] = '="Kumulatif s/d: "&TEXT(EOMONTH(DATE(TahunSel,BulanSel,1),0),"dd mmmm yyyy")'
    ws["B3"].font = SUBTITLE
    style_header(ws, 6, 1, ["KODE AKUN", "NAMA AKUN", "MUTASI DEBET", "MUTASI KREDIT", "SALDO DEBET", "SALDO KREDIT"])
    row = 7
    for code, name, _, _ in COA:
        ws.cell(row=row, column=1, value=code).number_format = "@"
        ws.cell(row=row, column=2, value=name)
        end_expr = "EOMONTH(DATE(TahunSel,BulanSel,1),0)"
        ws.cell(row=row, column=3, value=f'=SUMIFS({TX_DEBIT},{TX_CODE},$A{row},{TX_DATE},"<="&{end_expr})')
        ws.cell(row=row, column=4, value=f'=SUMIFS({TX_CREDIT},{TX_CODE},$A{row},{TX_DATE},"<="&{end_expr})')
        ws.cell(row=row, column=5, value=f"=IF(C{row}>D{row},C{row}-D{row},0)")
        ws.cell(row=row, column=6, value=f"=IF(D{row}>C{row},D{row}-C{row},0)")
        for col in range(3, 7):
            money(ws.cell(row=row, column=col))
        style_row(ws, row, 1, 6, fill=FILL_ALT if row % 2 == 0 else None)
        row += 1
    total_row = row
    ws.cell(row=row, column=2, value="TOTAL")
    for col in range(3, 7):
        letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f"=SUM({letter}7:{letter}{row - 1})")
        money(ws.cell(row=row, column=col))
    style_row(ws, row, 1, 6, fill=FILL_TOTAL, font=TOTAL_FONT)
    row += 2
    ws.cell(row=row, column=2, value="Laba/(Rugi) berjalan s/d periode")
    ws.cell(row=row, column=5, value=f"=IF('LABA RUGI'!$D${laba_rugi_rows['profit']}<0,ABS('LABA RUGI'!$D${laba_rugi_rows['profit']}),0)")
    ws.cell(row=row, column=6, value=f"=IF('LABA RUGI'!$D${laba_rugi_rows['profit']}>0,'LABA RUGI'!$D${laba_rugi_rows['profit']},0)")
    money(ws.cell(row=row, column=5))
    money(ws.cell(row=row, column=6))
    style_row(ws, row, 1, 6, fill=FILL_SECTION, font=TOTAL_FONT)
    row += 2
    check_row = row
    ws.cell(row=row, column=2, value="CHECK MUTASI DEBET = KREDIT")
    ws.cell(row=row, column=3, value=f'=IF(ABS(C{total_row}-D{total_row})<1,"OK — seimbang","CEK — tidak seimbang")')
    style_row(ws, row, 1, 6, fill=FILL_SECTION, font=TOTAL_FONT)
    ws.conditional_formatting.add(f"C{check_row}", FormulaRule(formula=[f'LEFT(C{check_row},2)="OK"'], fill=FILL_GREEN))
    ws.conditional_formatting.add(f"C{check_row}", FormulaRule(formula=[f'LEFT(C{check_row},2)<>"OK"'], fill=FILL_RED))
    ws.freeze_panes = "A7"
    return {"total": total_row, "check": check_row}


def build_posisi_keuangan(ws, laba_rugi_rows):
    ws.sheet_properties.tabColor = "C55A11"
    for col, width in {"A": 16, "B": 44, "C": 21, "D": 4, "E": 16, "F": 44, "G": 21}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "LAPORAN POSISI KEUANGAN (NERACA)"
    ws["B2"].font = TITLE
    ws["B3"] = '="Posisi per: "&TEXT(EOMONTH(DATE(TahunSel,BulanSel,1),0),"dd mmmm yyyy")'
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:G2")
    style_header(ws, 6, 1, ["KODE AKUN", "ASET", "JUMLAH", "", "KODE AKUN", "KEWAJIBAN & EKUITAS", "JUMLAH"])
    row = 7
    asset_start = row
    for code, name, _, _ in ASSETS:
        ws.cell(row=row, column=1, value=code).number_format = "@"
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=f'=SUMIFS({TX_DEBIT},{TX_CODE},$A{row},{TX_DATE},"<="&EOMONTH(DATE(TahunSel,BulanSel,1),0))-SUMIFS({TX_CREDIT},{TX_CODE},$A{row},{TX_DATE},"<="&EOMONTH(DATE(TahunSel,BulanSel,1),0))')
        money(ws.cell(row=row, column=3))
        style_row(ws, row, 1, 3, fill=FILL_ALT if row % 2 == 0 else None)
        row += 1
    asset_total = row
    ws.cell(row=row, column=2, value="TOTAL ASET")
    ws.cell(row=row, column=3, value=f"=SUM(C{asset_start}:C{row - 1})")
    money(ws.cell(row=row, column=3))
    style_row(ws, row, 1, 3, fill=FILL_TOTAL, font=TOTAL_FONT)

    right_row = 7
    right_start = right_row
    for code, name, _, _ in LIABILITIES + EQUITY:
        ws.cell(row=right_row, column=5, value=code).number_format = "@"
        ws.cell(row=right_row, column=6, value=name)
        ws.cell(row=right_row, column=7, value=f'=SUMIFS({TX_CREDIT},{TX_CODE},$E{right_row},{TX_DATE},"<="&EOMONTH(DATE(TahunSel,BulanSel,1),0))-SUMIFS({TX_DEBIT},{TX_CODE},$E{right_row},{TX_DATE},"<="&EOMONTH(DATE(TahunSel,BulanSel,1),0))')
        money(ws.cell(row=right_row, column=7))
        style_row(ws, right_row, 5, 7, fill=FILL_ALT if right_row % 2 == 0 else None)
        right_row += 1
    profit_row = right_row
    ws.cell(row=right_row, column=6, value="Laba/(Rugi) berjalan")
    ws.cell(row=right_row, column=7, value=f"='LABA RUGI'!$D${laba_rugi_rows['profit']}")
    money(ws.cell(right_row, column=7))
    style_row(ws, right_row, 5, 7, fill=FILL_SECTION, font=TOTAL_FONT)
    right_row += 1
    right_total = right_row
    ws.cell(row=right_row, column=6, value="TOTAL KEWAJIBAN & EKUITAS")
    ws.cell(row=right_row, column=7, value=f"=SUM(G{right_start}:G{right_row - 1})")
    money(ws.cell(right_row, column=7))
    style_row(ws, right_row, 5, 7, fill=FILL_TOTAL, font=TOTAL_FONT)

    diff_row = max(asset_total, right_total) + 2
    ws.cell(row=diff_row, column=2, value="SELISIH ASET - KEWAJIBAN & EKUITAS")
    ws.cell(row=diff_row, column=3, value=f"=C{asset_total}-G{right_total}")
    money(ws.cell(diff_row, column=3))
    style_row(ws, diff_row, 1, 7, fill=FILL_TOTAL, font=TOTAL_FONT)
    check_row = diff_row + 1
    ws.cell(row=check_row, column=2, value="CHECK BALANCE")
    ws.cell(row=check_row, column=3, value=f'=IF(ABS(C{diff_row})<1,"OK — balance","CEK — tidak balance")')
    style_row(ws, check_row, 1, 7, fill=FILL_SECTION, font=TOTAL_FONT)
    ws.conditional_formatting.add(f"C{check_row}", FormulaRule(formula=[f'LEFT(C{check_row},2)="OK"'], fill=FILL_GREEN))
    ws.conditional_formatting.add(f"C{check_row}", FormulaRule(formula=[f'LEFT(C{check_row},2)<>"OK"'], fill=FILL_RED))
    ws.freeze_panes = "A7"
    return {"asset_total": asset_total, "right_total": right_total, "diff": diff_row, "check": check_row}


def build_petunjuk(ws):
    ws.sheet_properties.tabColor = "7F7F7F"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 105
    ws["B2"] = "PETUNJUK PENGGUNAAN TEMPLATE"
    ws["B2"].font = TITLE
    ws.merge_cells("B2:B2")
    instructions = [
        (4, "1. Mulai dari PENGATURAN. Pastikan identitas, daftar unit, Bulan aktif, dan Tahun aktif sudah benar."),
        (5, "2. Input jurnal di TRANSAKSI. Isi tanggal, unit usaha, jenis, kode akun, uraian, serta hanya angka Debet atau Kredit pada setiap baris."),
        (6, "3. Satu transaksi dapat terdiri dari beberapa baris. Gunakan uraian yang konsisten, tetapi laporan tidak bergantung pada uraian: agregasi menggunakan Kode Akun dan tanggal."),
        (7, "4. Baris contoh September 2025 diberi awalan [CONTOH]. Hapus atau ganti baris tersebut sebelum dipakai sebagai laporan resmi."),
        (8, "5. Pastikan kontrol 'Jurnal Debet = Kredit' dan 'CHECK BALANCE' di PENGATURAN berstatus OK. Jika tidak, cari baris jurnal yang belum berpasangan."),
        (9, "6. Lihat BULANAN untuk tren Jan–Des, LABA RUGI untuk bulan aktif dan kumulatif, ARUS KAS untuk pergerakan rekening kas, serta NERACA SALDO dan POSISI KEUANGAN untuk pemeriksaan posisi."),
        (10, "7. Gunakan akun yang sudah tersedia di AKUN. Untuk akun baru permanen, tambahkan tuple akun ke COA di generate_template.py lalu jalankan ulang agar baris akun muncul di semua laporan; menambah baris AKUN saja tidak cukup."),
        (12, "IMPORT KE GOOGLE SHEETS"),
        (13, "Di Google Drive pilih New > File upload, unggah file XLSX ini, lalu Open with > Google Sheets. Alternatifnya di spreadsheet pilih File > Import > Upload dan pilih Create new spreadsheet."),
        (14, "Setelah impor, cek locale tanggal/rupiah, data validation di TRANSAKSI, serta jalankan ulang pemeriksaan kontrol. Google Sheets akan menghitung ulang formula saat file dibuka."),
        (16, "CATATAN DESAIN"),
        (17, "Template ini memakai satu tabel transaksi sepanjang tahun dan laporan berbasis SUMIFS dengan Kode Akun + periode. Hindari memindahkan/menyisipkan kolom di TRANSAKSI karena formula dan validasi mengikuti struktur tersebut."),
    ]
    for row, text in instructions:
        ws.cell(row=row, column=2, value=text)
        ws.cell(row=row, column=2).alignment = LEFT
        if row in (12, 16):
            ws.cell(row=row, column=2).font = Font(bold=True, color=SAND)
            ws.cell(row=row, column=2).fill = FILL_SECTION
        else:
            ws.cell(row=row, column=2).border = GRID
    ws.freeze_panes = "B4"


def main():
    wb = Workbook()
    first = wb.active
    first.title = "PENGATURAN"
    for sheet_name in ["TRANSAKSI", "BULANAN", "LABA RUGI", "ARUS KAS", "NERACA SALDO", "POSISI KEUANGAN", "AKUN", "PETUNJUK"]:
        wb.create_sheet(sheet_name)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    build_pengaturan(wb["PENGATURAN"])
    build_transaksi(wb["TRANSAKSI"], wb)
    build_akun(wb["AKUN"], wb)
    monthly_rows = build_bulanan(wb["BULANAN"])
    laba_rugi_rows = build_laba_rugi(wb["LABA RUGI"])
    arus_kas_rows = build_arus_kas(wb["ARUS KAS"])
    neraca_saldo_rows = build_neraca_saldo(wb["NERACA SALDO"], laba_rugi_rows)
    posisi_rows = build_posisi_keuangan(wb["POSISI KEUANGAN"], laba_rugi_rows)
    build_petunjuk(wb["PETUNJUK"])

    pg = wb["PENGATURAN"]
    pg["D21"] = "RINGKASAN BULAN AKTIF"
    pg["D21"].font = Font(bold=True, color=SAND)
    kpis = [
        (22, "Saldo kas akhir", f"='ARUS KAS'!$H${arus_kas_rows['total']}"),
        (23, "Pendapatan bulan ini", f"='LABA RUGI'!$C${laba_rugi_rows['income_total']}"),
        (24, "Beban bulan ini", f"='LABA RUGI'!$C${laba_rugi_rows['expense_total']}"),
        (25, "Surplus/(Rugi) bulan ini", f"='LABA RUGI'!$C${laba_rugi_rows['profit']}"),
        (26, "Surplus/(Rugi) s/d bulan ini", f"='LABA RUGI'!$D${laba_rugi_rows['profit']}"),
    ]
    for row, label, formula in kpis:
        pg.cell(row=row, column=4, value=label)
        pg.cell(row=row, column=5, value=formula)
        money(pg.cell(row=row, column=5))
        style_row(pg, row, 4, 5)
    pg["D28"] = "CEK KESEHATAN DATA"
    pg["D28"].font = Font(bold=True, color=SAND)
    checks = [
        (29, "Jurnal Debet = Kredit?", '=IF(ABS(SUM(TRANSAKSI!$H$10:$H$1000)-SUM(TRANSAKSI!$I$10:$I$1000))<1,"OK — seimbang","CEK — selisih")'),
        (30, "Neraca balance?", f"='POSISI KEUANGAN'!$C${posisi_rows['check']}"),
    ]
    for row, label, formula in checks:
        pg.cell(row=row, column=4, value=label)
        pg.cell(row=row, column=5, value=formula)
        style_row(pg, row, 4, 5)
        pg.conditional_formatting.add(f"E{row}", FormulaRule(formula=[f'LEFT(E{row},2)="OK"'], fill=FILL_GREEN))
        pg.conditional_formatting.add(f"E{row}", FormulaRule(formula=[f'LEFT(E{row},2)<>"OK"'], fill=FILL_RED))
    pg["D32"] = "Semua laporan otomatis mengikuti Bulan/Tahun aktif dan transaksi yang diisi."
    pg["D32"].font = SMALL
    pg.merge_cells("D32:G32")
    pg.freeze_panes = "A4"

    # Semua sheet dibuat dengan grid dan tampilan ringkas agar nyaman dicetak/impor.
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Generated: {OUTPUT}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"COA detail: {len(COA)} akun")


if __name__ == "__main__":
    main()
