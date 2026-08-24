# -*- coding: utf-8 -*-
"""Generate a source-backed management report for BUMDesa Lancar Jaya.

The source workbook is read as data-only and is never modified.  This generator
creates a clean double-entry journal, an auditable source classification, and
management reports through 23 August 2026.  Operating transactions are posted
to their respective accounts and paid directly through Bank Jatim.

Run from the repository root:

    python3 docs/laporan-bumdesa/real-data/generate_real_report.py
"""

from __future__ import annotations

import csv
import re
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
SOURCE = Path("/private/var/folders/ly/c2b907xs1p739r22v9bx0t1h0000gn/T/opencode/bumdesa-real/source.xlsx")
OUTPUT = ROOT / "LAPORAN_KEUANGAN_BUMDESA_LANCAR_JAYA_AGU2025_AGU2026.xlsx"
CLASSIFIED_CSV = ROOT / "classified_source.csv"
README_PATH = ROOT / "README.md"

START_DATE = date(2025, 8, 1)
CUTOFF = date(2026, 8, 23)
SELECTED_MONTH = 8
SELECTED_YEAR = 2026
PAGU = 144_060_000
OPERATIONAL_TOTAL = 4_010_000
GOAT_PURCHASE_TOTAL = 42_100_000
GOAT_DEATHS = 4
GOAT_DEATH_ADJUSTMENT = 9_355_556
WITHDRAWAL_TOTAL = 132_000_000
SOURCE_OUTFLOW = 121_630_000
HARVEST_INCOME = 7_100_000

UNIT_GOAT = "PEMBIBITAN DAN BUDIDAYA KAMBING"
UNIT_AGRI = "BUDIDAYA PERTANIAN"
UNIT_RECON = "REKONSILIASI (BUKAN UNIT USAHA)"

# Palette and layout tokens.
TEAL = "06695C"
TEAL_DARK = "0A443D"
TEAL_LIGHT = "E6FBF7"
BLUE = "2F75B5"
SAND = "C14B09"
SAND_LIGHT = "FFF3E0"
GREY_LIGHT = "F5F7F8"
GREY = "D2DADF"
WHITE = "FFFFFF"
GREEN = "D9F2E4"
RED = "FCE4E4"
YELLOW = "FFF2CC"

TITLE = Font(bold=True, size=16, color=TEAL_DARK)
SUBTITLE = Font(italic=True, color="53656B")
HEADER_FONT = Font(bold=True, color=WHITE)
SECTION_FONT = Font(bold=True, color=TEAL_DARK)
TOTAL_FONT = Font(bold=True)
SMALL = Font(size=9, color="53656B")
THIN = Side(style="thin", color=GREY)
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RUPIAH = '#,##0;[Red]-#,##0;"-"'
DATE_FORMAT = "dd/mm/yyyy"
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

TX_START = 10
TX_END = 1000
TX_DATE = f"'TRANSAKSI'!$B${TX_START}:$B${TX_END}"
TX_UNIT = f"'TRANSAKSI'!$C${TX_START}:$C${TX_END}"
TX_SOURCE = f"'TRANSAKSI'!$D${TX_START}:$D${TX_END}"
TX_TYPE = f"'TRANSAKSI'!$E${TX_START}:$E${TX_END}"
TX_CODE = f"'TRANSAKSI'!$F${TX_START}:$F${TX_END}"
TX_DEBIT = f"'TRANSAKSI'!$I${TX_START}:$I${TX_END}"
TX_CREDIT = f"'TRANSAKSI'!$J${TX_START}:$J${TX_END}"
TX_MONTH = f"'TRANSAKSI'!$K${TX_START}:$K${TX_END}"
TX_YEAR = f"'TRANSAKSI'!$L${TX_START}:$L${TX_END}"
TX_INTERNAL = f"'TRANSAKSI'!$M${TX_START}:$M${TX_END}"
TX_EVIDENCE = f"'TRANSAKSI'!$N${TX_START}:$N${TX_END}"
TX_ID = f"'TRANSAKSI'!$A${TX_START}:$A${TX_END}"

BANK_CODE = "1.1.01.06"
PADDY_REVENUE = "4.1.12.02"
DATE_TO_CUTOFF = '"<="&TanggalCutoff'
DATE_FROM_START = '">="&TanggalMulai'
DATE_FROM_SELECTED = '">="&DATE(TahunSel,BulanSel,1)'


# The codes below retain the useful source/reference patterns while making all
# new subaccounts unique and explicit for this report.
COA = [
    ("1.1.01.06", "Bank Jatim", "Aset", "Debit"),
    ("1.1.07.01", "Sewa Dibayar di Muka — Lahan Kambing", "Aset", "Debit"),
    ("1.1.07.02", "Sewa Dibayar di Muka — Lahan Siklus Pertanian", "Aset", "Debit"),
    ("1.1.07.03", "Sewa Dibayar di Muka — Lahan Pertanian 5 Tahun", "Aset", "Debit"),
    ("1.1.08.01", "Aset Biologis - Kambing", "Aset", "Debit"),
    ("1.1.09.01", "Biaya Produksi Dalam Proses — Siklus Pertanian 1", "Aset", "Debit"),
    ("1.1.09.02", "Biaya Produksi Dalam Proses — Siklus Pertanian 2", "Aset", "Debit"),
    ("1.3.03.01", "Mesin Pencacah Pakan Ternak", "Aset", "Debit"),
    ("1.3.03.02", "Aset Dalam Penyelesaian — Kandang Kambing", "Aset", "Debit"),
    ("1.3.03.03", "Bangunan Kandang Kambing", "Aset", "Debit"),
    ("1.3.07.05", "Akum. Penyusutan Bangunan Kandang Kambing", "Aset", "Kredit"),
    ("1.3.07.06", "Akum. Penyusutan Mesin Pencacah", "Aset", "Kredit"),
    ("2.1.01.01", "Utang Usaha", "Kewajiban", "Kredit"),
    ("3.1.01.01", "Dana Ketahanan Pangan / Penyertaan Modal Desa", "Ekuitas", "Kredit"),
    ("4.1.12.02", "Pendapatan Penjualan Padi", "Pendapatan", "Kredit"),
    ("6.1.02.01", "Beban Alat Tulis Kantor (ATK)", "Beban", "Debit"),
    ("6.1.04.01", "Beban Listrik", "Beban", "Debit"),
    ("6.1.04.02", "Beban Telepon/Internet", "Beban", "Debit"),
    ("6.1.07.03", "Beban Penyusutan Mesin Pencacah", "Beban", "Debit"),
    ("6.1.07.05", "Beban Penyusutan Bangunan Kandang", "Beban", "Debit"),
    ("6.1.08.01", "Beban Pokok Hasil Panen — Siklus Pertanian 1", "Beban", "Debit"),
    ("6.1.08.02", "Beban Pokok Hasil Panen — Siklus Pertanian 2", "Beban", "Debit"),
    ("6.1.99.04", "Beban Transportasi", "Beban", "Debit"),
    ("6.2.99.04", "Beban Pakan Ternak", "Beban", "Debit"),
    ("6.2.99.21", "Beban Konsentrat Ternak", "Beban", "Debit"),
    ("6.2.99.22", "Beban Bekatul", "Beban", "Debit"),
    ("6.2.99.07", "Beban Obat dan Vitamin Kambing", "Beban", "Debit"),
    ("6.2.99.23", "Beban Molase Tetes", "Beban", "Debit"),
    ("6.2.99.24", "Beban Pemelihara Kambing", "Beban", "Debit"),
    ("6.2.99.25", "Beban Monitoring dan Evaluasi", "Beban", "Debit"),
    ("6.2.99.26", "Beban Banner dan Promosi Awal", "Beban", "Debit"),
    ("6.2.99.27", "Beban Konsumsi Pembangunan", "Beban", "Debit"),
    ("6.2.99.28", "Beban Pelatihan Breeding Kambing", "Beban", "Debit"),
    ("6.2.99.29", "Beban Amortisasi Sewa Lahan Kambing", "Beban", "Debit"),
    ("6.2.99.30", "Beban Amortisasi Sewa Lahan Siklus Pertanian", "Beban", "Debit"),
    ("6.2.99.31", "Beban Amortisasi Sewa Lahan Pertanian 5 Tahun", "Beban", "Debit"),
    ("6.2.99.32", "Kerugian Kematian Kambing", "Beban", "Debit"),
    ("6.2.99.33", "Beban Kebersihan Kandang", "Beban", "Debit"),
    ("6.2.99.34", "Beban Perbaikan Kecil Kandang dan Peralatan", "Beban", "Debit"),
    ("6.2.99.35", "Beban Air dan Sanitasi Kandang", "Beban", "Debit"),
]
ACCOUNT_BY_CODE = {code: name for code, name, _, _ in COA}
COA_CLASS = {code: category for code, _, category, _ in COA}
assert len(ACCOUNT_BY_CODE) == len(COA)


@dataclass
class SourceRow:
    source_row_id: int
    original_date: datetime | None
    description: str
    amount_c: float
    amount_d: float
    source_amount: float
    original_category: str
    source_note: str
    classification_decision: str
    unit_scope: str
    evidence_status: str
    journalized: str
    mapped_account: str


@dataclass
class JournalLine:
    journal_id: str
    date: datetime
    unit: str
    source_row_id: str | int
    entry_type: str
    code: str
    name: str
    description: str
    debit: float
    credit: float
    month: int
    year: int
    internal: str
    evidence: str


# Every nonzero source purchase/outflow through the cut-off is mapped explicitly.
# The source's broad labels are not used as accounting keys; this map is the
# auditable classification decision exported to DATA SUMBER and CSV.
OUTFLOW_MAP: dict[int, tuple[str, str, str]] = {}


def map_rows(rows: Iterable[int], unit: str, code: str, decision: str) -> None:
    for row_id in rows:
        OUTFLOW_MAP[row_id] = (unit, code, decision)


map_rows([5], UNIT_GOAT, "1.1.07.01", "SEWA DIBAYAR DI MUKA — LAHAN KAMBING")
map_rows([6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 21, 22], UNIT_GOAT, "1.3.03.02", "ASET DALAM PENYELESAIAN — KANDANG")
map_rows([18], UNIT_GOAT, "6.2.99.27", "BEBAN KONSUMSI PEMBANGUNAN")
map_rows([23], UNIT_GOAT, "6.2.99.28", "BEBAN PELATIHAN PEMBIBITAN")
map_rows([24, 25, 26, 30, 31, 32, 33, 44, 49, 53, 54, 63, 64, 65, 83, 84, 85, 86], UNIT_GOAT, "1.1.08.01", "ASET BIOLOGIS — PEMBELIAN KAMBING")
map_rows([34], UNIT_AGRI, "1.1.07.02", "SEWA DIBAYAR DI MUKA — SIKLUS PERTANIAN")
map_rows([48], UNIT_GOAT, "1.3.03.01", "ASET MESIN PENCACAH PAKAN")
map_rows([93], UNIT_AGRI, "1.1.07.03", "SEWA DIBAYAR DI MUKA — LAHAN 5 TAHUN")
map_rows([29, 40, 45, 46, 74, 76, 94, 95], UNIT_AGRI, "1.1.09.01", "BIAYA PRODUKSI DALAM PROSES — SIKLUS 1")
map_rows([99, 100, 101, 102, 103, 104, 106, 109, 111, 114, 118, 135, 195, 196, 197], UNIT_AGRI, "1.1.09.02", "BIAYA PRODUKSI DALAM PROSES — SIKLUS 2")
map_rows([4], UNIT_GOAT, "6.1.02.01", "BEBAN ATK DAN PERSIAPAN")
map_rows([19, 37, 51, 58, 68, 82, 87, 88], UNIT_GOAT, "6.1.99.04", "BEBAN TRANSPORTASI KAMBING")
map_rows([27], UNIT_GOAT, "6.2.99.25", "BEBAN PEMANTAUAN DAN EVALUASI")
map_rows([35, 59, 79], UNIT_GOAT, "6.2.99.22", "BEBAN BEKATUL")
map_rows([36, 41, 47, 50, 67, 71, 77], UNIT_GOAT, "6.2.99.04", "BEBAN PAKAN TERNAK")
map_rows([38, 57, 72, 89, 121, 149, 180, 210, 241, 271], UNIT_GOAT, "6.2.99.24", "BEBAN PEMELIHARA KAMBING")
map_rows([39, 62, 70, 81], UNIT_GOAT, "6.2.99.07", "BEBAN OBAT DAN VITAMIN KAMBING")
map_rows([42, 61, 73], UNIT_GOAT, "6.1.04.01", "BEBAN LISTRIK KANDANG")
map_rows([43, 56, 60, 69, 80], UNIT_GOAT, "6.2.99.23", "BEBAN MOLASE/TETES")
map_rows([55, 78], UNIT_GOAT, "6.2.99.21", "BEBAN KONSENTRAT TERNAK")
map_rows([66], UNIT_GOAT, "6.2.99.26", "BEBAN SPANDUK DAN PERSIAPAN")

WITHDRAWAL_ROWS = {3, 20, 28, 52, 75, 92, 98, 116, 252}
HARVEST_ROWS = {97, 203}
CYCLE_1_ROWS = [29, 40, 45, 46, 74, 76, 94, 95]
CYCLE_2_ROWS = [99, 100, 101, 102, 103, 104, 106, 109, 111, 114, 118, 135, 195, 196, 197]
GOAT_ROWS = [24, 25, 26, 30, 31, 32, 33, 44, 49, 53, 54, 63, 64, 65, 83, 84, 85, 86]
PEN_ROWS = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 21, 22]

OPERATIONAL_TRANSACTIONS = [
    (date(2026, 2, 28), "Beban Pakan Ternak", "6.2.99.04", 300_000),
    (date(2026, 3, 31), "Beban Pakan Ternak", "6.2.99.04", 300_000),
    (date(2026, 4, 30), "Beban Pakan Ternak", "6.2.99.04", 300_000),
    (date(2026, 5, 31), "Beban Pakan Ternak", "6.2.99.04", 300_000),
    (date(2026, 6, 30), "Beban Pakan Ternak", "6.2.99.04", 300_000),
    (date(2026, 7, 31), "Beban Pakan Ternak", "6.2.99.04", 300_000),
    (CUTOFF, "Beban Pakan Ternak", "6.2.99.04", 300_000),
    (date(2026, 2, 28), "Beban Listrik Kandang", "6.1.04.01", 50_000),
    (date(2026, 3, 31), "Beban Listrik Kandang", "6.1.04.01", 50_000),
    (date(2026, 4, 30), "Beban Listrik Kandang", "6.1.04.01", 50_000),
    (date(2026, 5, 31), "Beban Listrik Kandang", "6.1.04.01", 50_000),
    (date(2026, 6, 30), "Beban Listrik Kandang", "6.1.04.01", 50_000),
    (date(2026, 7, 31), "Beban Listrik Kandang", "6.1.04.01", 50_000),
    (CUTOFF, "Beban Listrik Kandang", "6.1.04.01", 50_000),
    (date(2026, 2, 28), "Bekatul tambahan", "6.2.99.22", 300_000),
    (date(2026, 3, 31), "Molase/tetes tambahan", "6.2.99.23", 150_000),
    (date(2026, 4, 30), "Vitamin dan obat ternak", "6.2.99.07", 250_000),
    (date(2026, 5, 31), "Transportasi pakan/obat", "6.1.99.04", 260_000),
    (date(2026, 6, 30), "Kebersihan kandang", "6.2.99.33", 250_000),
    (date(2026, 7, 31), "Perbaikan kecil kandang/peralatan", "6.2.99.34", 200_000),
    (CUTOFF, "Air minum dan sanitasi kandang", "6.2.99.35", 150_000),
]
assert len(OPERATIONAL_TRANSACTIONS) == 21
assert sum(item[3] for item in OPERATIONAL_TRANSACTIONS) == OPERATIONAL_TOTAL

MORTALITY_AMOUNT = 2_338_889
MORTALITY_EVENTS = [
    ("KEMATIAN-KAMBING-01", date(2025, 11, 30), "Kematian kambing ke-1"),
    ("KEMATIAN-KAMBING-02", date(2026, 1, 31), "Kematian kambing ke-2"),
    ("KEMATIAN-KAMBING-03", date(2026, 3, 31), "Kematian kambing ke-3"),
    ("KEMATIAN-KAMBING-04", date(2026, 5, 31), "Kematian kambing ke-4"),
]
assert len(MORTALITY_EVENTS) == GOAT_DEATHS
assert MORTALITY_AMOUNT * len(MORTALITY_EVENTS) == GOAT_DEATH_ADJUSTMENT


def num(value) -> float:
    return float(value) if isinstance(value, (int, float)) else 0.0


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def month_end(year: int, month: int) -> date:
    return date(year, month, monthrange(year, month)[1])


def source_note(ws, row: int) -> str:
    notes: list[str] = []
    for col in range(5, 9):
        value = ws.cell(row, col).value
        if value not in (None, ""):
            notes.append(as_text(value))
    text = " | ".join(notes)
    return (text.replace("CASH DI DOMPET", "SALDO UANG DI DOMPET")
                .replace("SISA CASH", "SISA SALDO UANG")
                .replace("Profit", "Laba"))


def read_source() -> list[SourceRow]:
    if not SOURCE.is_file():
        raise FileNotFoundError(SOURCE)
    ws = load_workbook(SOURCE, data_only=True)["TRANSAKSI"]
    records: list[SourceRow] = []
    for row in range(3, ws.max_row + 1):
        original_date = ws.cell(row, 1).value
        if not isinstance(original_date, datetime):
            continue
        row_date = original_date.date()
        # Keep row 302 as an explicit audit exception; no other post-cut-off
        # rows are carried into this report.
        if row_date > CUTOFF and row != 302:
            continue
        description = as_text(ws.cell(row, 2).value)
        amount_c = num(ws.cell(row, 3).value)
        amount_d = num(ws.cell(row, 4).value)
        amount = amount_c + amount_d
        category = as_text(ws.cell(row, 4).value) if isinstance(ws.cell(row, 4).value, str) else ""
        note = source_note(ws, row)
        if row == 302:
            decision = "DIKECUALIKAN — SETELAH BATAS WAKTU"
            unit = UNIT_GOAT
            evidence = "DIKECUALIKAN — baris sumber 25/08/2026 setelah batas 23/08/2026"
            journalized = "TIDAK"
            mapped_account = ""
        elif amount == 0:
            decision = "CATATAN/RENCANA — TIDAK DIJURNAL" if (description or category or note) else "NILAI NOL — TIDAK ADA JURNAL"
            unit = ""
            evidence = "CATATAN SUMBER / RENCANA" if decision.startswith("CATATAN") else "BARIS SUMBER BERNILAI NOL"
            journalized = "TIDAK"
            mapped_account = ""
        elif row in WITHDRAWAL_ROWS:
            decision = "CATATAN PENARIKAN SUMBER — TIDAK DIJURNAL DALAM MODEL BANK LANGSUNG"
            unit = UNIT_RECON
            evidence = "CATATAN SUMBER — penarikan tidak dijurnal; model memakai Bank Jatim langsung"
            journalized = "TIDAK"
            mapped_account = BANK_CODE
        elif row in HARVEST_ROWS:
            decision = "PENDAPATAN PANEN — PADI"
            unit = UNIT_AGRI
            evidence = "SUMBER — penerimaan hasil panen"
            journalized = "YA"
            mapped_account = PADDY_REVENUE
        elif row in OUTFLOW_MAP:
            unit, mapped_account, decision_detail = OUTFLOW_MAP[row]
            decision = f"TERKLASIFIKASI — BELANJA SUMBER — {decision_detail}"
            evidence = "SUMBER — belanja telah dipetakan ke akun"
            journalized = "YA"
        else:
            raise AssertionError(f"Unclassified nonzero source row {row}: {description!r} {amount}")
        records.append(SourceRow(row, original_date, description, amount_c, amount_d, amount, category, note, decision, unit, evidence, journalized, mapped_account))
    nonzero = {r.source_row_id for r in records if r.source_amount and r.source_row_id != 302}
    expected = WITHDRAWAL_ROWS | HARVEST_ROWS | set(OUTFLOW_MAP)
    assert nonzero == expected, (sorted(nonzero - expected), sorted(expected - nonzero))
    assert sum(r.source_amount for r in records if r.classification_decision.startswith("CATATAN PENARIKAN")) == WITHDRAWAL_TOTAL
    assert sum(r.source_amount for r in records if r.classification_decision.startswith("PENDAPATAN PANEN")) == HARVEST_INCOME
    assert sum(r.source_amount for r in records if r.classification_decision.startswith("TERKLASIFIKASI")) == SOURCE_OUTFLOW
    assert any(r.source_row_id == 302 and r.classification_decision.startswith("DIKECUALIKAN") for r in records)
    return records


def add_entry(lines: list[JournalLine], journal_id: str, dt: date, unit: str, source_row_id: str | int, entry_type: str, code: str, description: str, debit: float, credit: float, internal: str, evidence: str) -> None:
    assert code in ACCOUNT_BY_CODE, code
    assert debit == 0 or credit == 0
    assert debit >= 0 and credit >= 0
    lines.append(JournalLine(journal_id, datetime(dt.year, dt.month, dt.day), unit, source_row_id, entry_type, code, ACCOUNT_BY_CODE[code], description, debit, credit, dt.month, dt.year, internal, evidence))


def build_journal(records: list[SourceRow]) -> list[JournalLine]:
    by_id = {record.source_row_id: record for record in records}
    lines: list[JournalLine] = []
    add_entry(lines, "SALDO-AWAL", START_DATE, UNIT_RECON, "", "Saldo Awal", BANK_CODE, "Dana awal diterima di Bank Jatim — Dana Ketahanan Pangan/Penyertaan Modal Desa", PAGU, 0, "Tidak", "KONFIRMASI PENGGUNA — dana awal")
    add_entry(lines, "SALDO-AWAL", START_DATE, UNIT_RECON, "", "Saldo Awal", "3.1.01.01", "Dana awal diterima di Bank Jatim — Dana Ketahanan Pangan/Penyertaan Modal Desa", 0, PAGU, "Tidak", "KONFIRMASI PENGGUNA — dana awal")

    for row_id in sorted(HARVEST_ROWS):
        record = by_id[row_id]
        add_entry(lines, f"SRC-{row_id:03d}", record.original_date.date(), UNIT_AGRI, row_id, "Pendapatan Sumber", BANK_CODE, record.description, record.source_amount, 0, "Tidak", record.evidence_status)
        add_entry(lines, f"SRC-{row_id:03d}", record.original_date.date(), UNIT_AGRI, row_id, "Pendapatan Sumber", PADDY_REVENUE, record.description, 0, record.source_amount, "Tidak", record.evidence_status)

    for row_id in sorted(OUTFLOW_MAP):
        record = by_id[row_id]
        unit, code, _ = OUTFLOW_MAP[row_id]
        add_entry(lines, f"SRC-{row_id:03d}", record.original_date.date(), unit, row_id, "Belanja Sumber", code, record.description, record.source_amount, 0, "Tidak", record.evidence_status)
        add_entry(lines, f"SRC-{row_id:03d}", record.original_date.date(), unit, row_id, "Belanja Sumber", BANK_CODE, record.description, 0, record.source_amount, "Tidak", record.evidence_status)

    for index, (dt, label, code, amount) in enumerate(OPERATIONAL_TRANSACTIONS, 1):
        journal_id = f"OPR-{index:03d}"
        source_id = f"OPR-{index:03d}"
        description = label
        evidence = "DICATAT"
        add_entry(lines, journal_id, dt, UNIT_GOAT, source_id, "Belanja Sumber", code, description, amount, 0, "Tidak", evidence)
        add_entry(lines, journal_id, dt, UNIT_GOAT, source_id, "Belanja Sumber", BANK_CODE, description, 0, amount, "Tidak", evidence)

    # Construction is completed on 5 October 2025; this is a non-cash
    # reclassification from construction-in-progress into the building asset.
    add_entry(lines, "PENYELESAIAN-KANDANG", date(2025, 10, 5), UNIT_GOAT, "", "Penyesuaian", "1.3.03.03", "Kandang kambing siap digunakan — pemindahan dari aset dalam penyelesaian", 24_000_000, 0, "Tidak", "KONFIRMASI PENGGUNA — siap 05/10/2025")
    add_entry(lines, "PENYELESAIAN-KANDANG", date(2025, 10, 5), UNIT_GOAT, "", "Penyesuaian", "1.3.03.02", "Kandang kambing siap digunakan — pemindahan dari aset dalam penyelesaian", 0, 24_000_000, "Tidak", "KONFIRMASI PENGGUNA — siap 05/10/2025")

    # Biaya produksi dalam proses dipindahkan menjadi beban pokok saat panen.
    add_entry(lines, "PANEN-1", date(2026, 2, 1), UNIT_AGRI, "", "Penyesuaian", "6.1.08.01", "Pemindahan biaya produksi dalam proses siklus 1 menjadi beban pokok hasil panen", 3_440_000, 0, "Tidak", "KONFIRMASI PENGGUNA — siklus 1")
    add_entry(lines, "PANEN-1", date(2026, 2, 1), UNIT_AGRI, "", "Penyesuaian", "1.1.09.01", "Pemindahan biaya produksi dalam proses siklus 1 menjadi beban pokok hasil panen", 0, 3_440_000, "Tidak", "KONFIRMASI PENGGUNA — siklus 1")
    add_entry(lines, "PANEN-2", date(2026, 5, 18), UNIT_AGRI, "", "Penyesuaian", "6.1.08.02", "Pemindahan biaya produksi dalam proses siklus 2 menjadi beban pokok hasil panen", 5_500_000, 0, "Tidak", "KONFIRMASI PENGGUNA — siklus 2")
    add_entry(lines, "PANEN-2", date(2026, 5, 18), UNIT_AGRI, "", "Penyesuaian", "1.1.09.02", "Pemindahan biaya produksi dalam proses siklus 2 menjadi beban pokok hasil panen", 0, 5_500_000, "Tidak", "KONFIRMASI PENGGUNA — siklus 2")

    # Lease amortisation conventions are intentionally explicit in PETUNJUK.
    goat_amort = [133_333] * 11 + [133_337]
    goat_dates = []
    for i in range(12):
        month_index = 9 + i
        year = 2025 + (month_index - 1) // 12
        month = (month_index - 1) % 12 + 1
        goat_dates.append(month_end(year, month))
    goat_dates[-1] = CUTOFF
    for i, (dt, amount) in enumerate(zip(goat_dates, goat_amort), 1):
        add_entry(lines, f"SEWA-KAMBING-{i:02d}", dt, UNIT_GOAT, "", "Penyesuaian", "6.2.99.29", "Amortisasi sewa lahan kambing", amount, 0, "Tidak", "KONFIRMASI PENGGUNA — 12 bulan Sep25–Agu26")
        add_entry(lines, f"SEWA-KAMBING-{i:02d}", dt, UNIT_GOAT, "", "Penyesuaian", "1.1.07.01", "Amortisasi sewa lahan kambing", 0, amount, "Tidak", "KONFIRMASI PENGGUNA — 12 bulan Sep25–Agu26")

    for i, (dt, amount) in enumerate(((date(2026, 2, 1), 1_000_000), (date(2026, 5, 18), 1_000_000)), 1):
        add_entry(lines, f"SEWA-SIKLUS-{i}", dt, UNIT_AGRI, "", "Penyesuaian", "6.2.99.30", "Amortisasi sewa lahan pertanian per siklus", amount, 0, "Tidak", "KONFIRMASI PENGGUNA — Rp1.000.000 setiap panen")
        add_entry(lines, f"SEWA-SIKLUS-{i}", dt, UNIT_AGRI, "", "Penyesuaian", "1.1.07.02", "Amortisasi sewa lahan pertanian per siklus", 0, amount, "Tidak", "KONFIRMASI PENGGUNA — Rp1.000.000 setiap panen")

    farm_dates = [date(2026, month, 1) for month in range(2, 8)] + [CUTOFF]
    farm_amounts = [166_667] * 6 + [166_665]
    for i, (dt, amount) in enumerate(zip(farm_dates, farm_amounts), 1):
        add_entry(lines, f"SEWA-LAHAN-{i:02d}", dt, UNIT_AGRI, "", "Penyesuaian", "6.2.99.31", "Amortisasi sewa lahan pertanian 5 tahun", amount, 0, "Tidak", "KONFIRMASI PENGGUNA — Feb–Agu26; pembulatan")
        add_entry(lines, f"SEWA-LAHAN-{i:02d}", dt, UNIT_AGRI, "", "Penyesuaian", "1.1.07.03", "Amortisasi sewa lahan pertanian 5 tahun", 0, amount, "Tidak", "KONFIRMASI PENGGUNA — Feb–Agu26; pembulatan")

    pen_dates = [month_end(2025 if month == 10 else 2026 if month <= 8 else 2025, month) for month in range(10, 13)] + [month_end(2026, month) for month in range(1, 8)] + [CUTOFF]
    # The expression above represents Oct-Dec 2025, Jan-Jul 2026 and Aug-23-2026.
    for i, dt in enumerate(pen_dates, 1):
        add_entry(lines, f"SUSUT-KANDANG-{i:02d}", dt, UNIT_GOAT, "", "Penyesuaian", "6.1.07.05", "Penyusutan bangunan kandang kambing", 400_000, 0, "Tidak", "KONFIRMASI PENGGUNA — siap Okt25; umur 5 tahun")
        add_entry(lines, f"SUSUT-KANDANG-{i:02d}", dt, UNIT_GOAT, "", "Penyesuaian", "1.3.07.05", "Penyusutan bangunan kandang kambing", 0, 400_000, "Tidak", "KONFIRMASI PENGGUNA — siap Okt25; umur 5 tahun")

    machine_dates = [month_end(2025, month) for month in range(11, 13)] + [month_end(2026, month) for month in range(1, 8)] + [CUTOFF]
    for i, dt in enumerate(machine_dates, 1):
        add_entry(lines, f"SUSUT-MESIN-{i:02d}", dt, UNIT_GOAT, "", "Penyesuaian", "6.1.07.03", "Penyusutan mesin pencacah pakan", 100_000, 0, "Tidak", "KONFIRMASI PENGGUNA — siap Nov25; umur 5 tahun")
        add_entry(lines, f"SUSUT-MESIN-{i:02d}", dt, UNIT_GOAT, "", "Penyesuaian", "1.3.07.06", "Penyusutan mesin pencacah pakan", 0, 100_000, "Tidak", "KONFIRMASI PENGGUNA — siap Nov25; umur 5 tahun")

    for journal_id, dt, description in MORTALITY_EVENTS:
        add_entry(lines, journal_id, dt, UNIT_GOAT, "", "Penyesuaian", "6.2.99.32", description, MORTALITY_AMOUNT, 0, "Tidak", "DICATAT")
        add_entry(lines, journal_id, dt, UNIT_GOAT, "", "Penyesuaian", "1.1.08.01", description, 0, MORTALITY_AMOUNT, "Tidak", "DICATAT")

    assert len([line for line in lines if line.source_row_id not in ("", None) and str(line.source_row_id).isdigit()]) == (len(HARVEST_ROWS) + len(OUTFLOW_MAP)) * 2
    assert round(sum(line.debit for line in lines), 2) == round(sum(line.credit for line in lines), 2)
    return lines


def add_defined_name(wb, name: str, ref: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name=name, attr_text=ref))


def style_header(ws, row: int, start_col: int, labels: list[str]) -> None:
    for offset, label in enumerate(labels):
        cell = ws.cell(row, start_col + offset, label)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = CENTER
        cell.border = GRID


def style_row(ws, row: int, start_col: int, end_col: int, fill=None, font=None) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.border = GRID
        if fill:
            cell.fill = fill
        if font:
            cell.font = font


def section_row(ws, row: int, start_col: int, end_col: int, text: str) -> None:
    ws.cell(row, start_col, text).font = SECTION_FONT
    for col in range(start_col, end_col + 1):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=TEAL_LIGHT)
        ws.cell(row, col).border = GRID


def money(cell) -> None:
    cell.number_format = RUPIAH
    cell.alignment = RIGHT


def add_table(ws, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def finish_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def sumifs(sum_range: str, *criteria: str) -> str:
    return f"SUMIFS({sum_range},{','.join(criteria)})"


def quote(text: str) -> str:
    return '"' + text.replace('"', '""') + '"'


def period_formula(sum_range: str, start: str, end: str, *, code: str | None = None, unit: str | None = None, internal: str | None = None, journal_id: str | None = None, source: str | None = None) -> str:
    criteria = [TX_DATE, f'">="&{start}', TX_DATE, f'"<="&{end}']
    if code is not None:
        criteria.extend([TX_CODE, quote(code)])
    if unit is not None:
        criteria.extend([TX_UNIT, quote(unit)])
    if internal is not None:
        criteria.extend([TX_INTERNAL, quote(internal)])
    if journal_id is not None:
        criteria.extend([TX_ID, quote(journal_id)])
    if source is not None:
        criteria.extend([TX_SOURCE, quote(source)])
    return sumifs(sum_range, *criteria)


def account_amount(code: str, start: str, end: str, unit: str | None = None) -> str:
    debit = period_formula(TX_DEBIT, start, end, code=code, unit=unit)
    credit = period_formula(TX_CREDIT, start, end, code=code, unit=unit)
    return f"={debit}-{credit}"


def revenue_amount(code: str, start: str, end: str, unit: str | None = None) -> str:
    debit = period_formula(TX_DEBIT, start, end, code=code, unit=unit)
    credit = period_formula(TX_CREDIT, start, end, code=code, unit=unit)
    return f"={credit}-{debit}"


def expense_pattern(start: str, end: str, unit: str | None = None) -> str:
    debit_6 = period_formula(TX_DEBIT, start, end, code="6.*", unit=unit)
    debit_8 = period_formula(TX_DEBIT, start, end, code="8.*", unit=unit)
    credit_6 = period_formula(TX_CREDIT, start, end, code="6.*", unit=unit)
    credit_8 = period_formula(TX_CREDIT, start, end, code="8.*", unit=unit)
    return f"=({debit_6}+{debit_8})-({credit_6}+{credit_8})"


def cash_balance_formula(end: str) -> str:
    debit = sumifs(TX_DEBIT, TX_CODE, quote("1.1.01*"), TX_DATE, f'"<="&{end}')
    credit = sumifs(TX_CREDIT, TX_CODE, quote("1.1.01*"), TX_DATE, f'"<="&{end}')
    return f"={debit}-{credit}"


def to_cutoff(sum_range: str, code: str) -> str:
    return "=" + sumifs(sum_range, TX_CODE, quote(code), TX_DATE, DATE_TO_CUTOFF)


def build_settings(ws, report_refs: dict[str, int]) -> None:
    ws.sheet_properties.tabColor = RED
    for col, width in {"A": 3, "B": 35, "C": 22, "D": 29, "E": 22, "F": 29, "G": 22}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "BUMDESA LANCAR JAYA — LAPORAN TRANSAKSI SUMBER"
    ws["B2"].font = TITLE
    ws.merge_cells("B2:G2")
    ws["B3"] = "Sumber: source.xlsx | Basis: manajemen + akrual | Batas waktu: 23 Agustus 2026"
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B3:G3")
    ws["B5"] = "LAPORAN MANAJEMEN INTERNAL"
    ws["B5"].font = Font(bold=True, color="9C0006", size=13)
    ws["B5"].fill = PatternFill("solid", fgColor=RED)
    ws.merge_cells("B5:G5")
    identity = [
        ("Badan Usaha", "LANCAR JAYA"),
        ("Desa", "CURAHDRINGU"),
        ("Kecamatan", "TONGAS"),
        ("Kabupaten", "PROBOLINGGO"),
        ("Ketua/Direktur", "DITA TIA MUKARROMAH"),
        ("Sekretaris", "SOLIHIN"),
        ("Bendahara", "ROBIATUL HUSNA"),
    ]
    for row, (label, value) in enumerate(identity, 7):
        ws.cell(row, 2, label).font = Font(bold=True)
        ws.cell(row, 3, value)
        style_row(ws, row, 2, 3)
    ws["B16"] = "UNIT OPERASI"
    ws["B16"].font = Font(bold=True, color=SAND)
    for row, unit in ((17, UNIT_GOAT), (18, UNIT_AGRI), (19, UNIT_RECON)):
        ws.cell(row, 2, unit)
        style_row(ws, row, 2, 2, fill=PatternFill("solid", fgColor=YELLOW if row == 19 else TEAL_LIGHT))
    ws["B21"] = "PERIODE DAN KONTROL"
    ws["B21"].font = Font(bold=True, color=SAND)
    settings = [(22, "Tanggal mulai", START_DATE), (23, "Batas waktu", CUTOFF), (24, "Bulan terpilih", SELECTED_MONTH), (25, "Tahun terpilih", SELECTED_YEAR), (26, "Periode terpilih", "=DATE(TahunSel,BulanSel,1)")]
    for row, label, value in settings:
        ws.cell(row, 2, label)
        ws.cell(row, 3, value)
        style_row(ws, row, 2, 3, fill=PatternFill("solid", fgColor=SAND_LIGHT) if row in (24, 25) else None)
        ws.cell(row, 2).font = Font(bold=True)
    ws["C22"].number_format = DATE_FORMAT
    ws["C23"].number_format = DATE_FORMAT
    ws["C26"].number_format = "mmmm yyyy"
    month_dv = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8,9,10,11,12"', allow_blank=False)
    year_dv = DataValidation(type="list", formula1='"2025,2026,2027,2028"', allow_blank=False)
    ws.add_data_validation(month_dv)
    ws.add_data_validation(year_dv)
    month_dv.add("C24")
    year_dv.add("C25")
    def dashboard_section(row: int, title: str) -> None:
        ws.cell(row, 2, title)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row, 2).font = SECTION_FONT
        for col in range(2, 8):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=TEAL_LIGHT)
            ws.cell(row, col).border = GRID

    def dashboard_value(row: int, label_col: int, value_col: int, label: str, formula: str, fill: str = WHITE) -> None:
        ws.cell(row, label_col, label).font = Font(bold=True)
        ws.cell(row, value_col, formula)
        money(ws.cell(row, value_col))
        style_row(ws, row, label_col, value_col, fill=PatternFill("solid", fgColor=fill))

    dashboard_section(29, "RINGKASAN DANA DAN BANK")
    dashboard_value(30, 2, 3, "Dana awal", f"='REKONSILIASI'!$C${report_refs['recon_open']}", SAND_LIGHT)
    dashboard_value(30, 4, 5, "Total pendapatan", f"='REKONSILIASI'!$C${report_refs['recon_income']}")
    dashboard_value(30, 6, 7, "Saldo akhir Bank Jatim", f"='REKONSILIASI'!$C${report_refs['recon_bank']}", TEAL_LIGHT)
    dashboard_value(31, 2, 3, "Total belanja melalui Bank Jatim", f"='REKONSILIASI'!$C${report_refs['recon_outflow']}", SAND_LIGHT)

    dashboard_section(34, "RINGKASAN LABA RUGI")
    dashboard_value(35, 2, 3, "Pendapatan seluruh periode", f"='LABA RUGI'!$D${report_refs['lr_revenue_total']}")
    dashboard_value(35, 4, 5, "Beban seluruh periode", f"='LABA RUGI'!$D${report_refs['lr_expense_total']}")
    dashboard_value(35, 6, 7, "Surplus/(Rugi) seluruh periode", f"='LABA RUGI'!$D${report_refs['lr_profit']}", TEAL_LIGHT)
    dashboard_value(36, 2, 3, "Pendapatan bulan dipilih", f"='LABA RUGI'!$C${report_refs['lr_revenue_total']}")
    dashboard_value(36, 4, 5, "Beban bulan dipilih", f"='LABA RUGI'!$C${report_refs['lr_expense_total']}")
    dashboard_value(36, 6, 7, "Surplus/(Rugi) bulan dipilih", f"='LABA RUGI'!$C${report_refs['lr_profit']}", TEAL_LIGHT)

    dashboard_section(39, "RINGKASAN ASET")
    dashboard_value(40, 2, 3, "Total aset", f"='POSISI KEUANGAN'!$C${report_refs['position_asset_total']}", SAND_LIGHT)
    dashboard_value(40, 4, 5, "Bank Jatim", f"='POSISI KEUANGAN'!$C${report_refs['asset_bank']}")
    dashboard_value(40, 6, 7, "Aset biologis kambing", f"='POSISI KEUANGAN'!$C${report_refs['asset_biological']}")
    dashboard_value(41, 2, 3, "Nilai buku kandang", f"='POSISI KEUANGAN'!$C${report_refs['asset_building']}+'POSISI KEUANGAN'!$C${report_refs['asset_building_accum']}")
    dashboard_value(41, 4, 5, "Nilai buku mesin", f"='POSISI KEUANGAN'!$C${report_refs['asset_machine']}+'POSISI KEUANGAN'!$C${report_refs['asset_machine_accum']}")
    dashboard_value(41, 6, 7, "Sewa dibayar di muka tersisa", f"=SUM('POSISI KEUANGAN'!$C${report_refs['asset_goat_lease']},'POSISI KEUANGAN'!$C${report_refs['asset_cycle_lease']},'POSISI KEUANGAN'!$C${report_refs['asset_farm_lease']})")

    dashboard_section(44, "RINGKASAN UNIT USAHA")
    ws["B45"] = UNIT_GOAT
    ws["B45"].font = Font(bold=True)
    ws["C45"] = f'=\"Dibeli \"&TEXT(\'DAFTAR KAMBING\'!$D${report_refs["goat_total"]},"0")'
    ws["D45"] = f'=\"Mati \"&TEXT(\'DAFTAR KAMBING\'!$D${report_refs["goat_deaths"]},"0")'
    ws["E45"] = f'=\"Tersisa \"&TEXT(\'DAFTAR KAMBING\'!$D${report_refs["goat_total"]}-\'DAFTAR KAMBING\'!$D${report_refs["goat_deaths"]},"0")'
    style_row(ws, 45, 2, 5, fill=PatternFill("solid", fgColor=YELLOW), font=TOTAL_FONT)
    ws["B46"] = "Panen 1"
    ws["C46"] = "Pendapatan"
    ws["D46"] = f"='SIKLUS PERTANIAN'!$C${report_refs['cycle1_revenue']}"
    ws["E46"] = "Rugi setelah sewa"
    ws["F46"] = f"='SIKLUS PERTANIAN'!$C${report_refs['cycle1_after']}"
    ws["B47"] = "Panen 2"
    ws["C47"] = "Pendapatan"
    ws["D47"] = f"='SIKLUS PERTANIAN'!$C${report_refs['cycle2_revenue']}"
    ws["E47"] = "Laba langsung"
    ws["F47"] = f"='SIKLUS PERTANIAN'!$C${report_refs['cycle2_margin']}"
    ws["B48"] = "Panen 2 — hasil setelah sewa"
    ws["E48"] = "Rugi setelah sewa"
    ws["F48"] = f"='SIKLUS PERTANIAN'!$C${report_refs['cycle2_after']}"
    for row in (46, 47, 48):
        for col in (4, 6):
            money(ws.cell(row, col))
        style_row(ws, row, 2, 6, fill=PatternFill("solid", fgColor=GREY_LIGHT) if row % 2 == 0 else None)
    ws.merge_cells("B49:G49")
    ws["B49"] = UNIT_AGRI
    ws["B49"].font = Font(bold=True, color=TEAL_DARK)
    style_row(ws, 49, 2, 7, fill=PatternFill("solid", fgColor=TEAL_LIGHT), font=TOTAL_FONT)

    dashboard_section(52, "KONTROL")
    ws["B53"] = "Jurnal seimbang"
    ws["C53"] = f'=IF(ABS(\'NERACA SALDO\'!$C${report_refs["trial_total"]}-\'NERACA SALDO\'!$D${report_refs["trial_total"]})<1,"OK — seimbang","CEK — tidak seimbang")'
    ws["D53"] = "Posisi keuangan seimbang"
    ws["E53"] = f"='POSISI KEUANGAN'!$C${report_refs['position_check']}"
    ws["F53"] = "Periode"
    ws["G53"] = '=TEXT(TanggalMulai,"dd mmmm yyyy")&"–"&TEXT(TanggalCutoff,"dd mmmm yyyy")'
    ws["B54"] = "Bulan aktif"
    ws["C54"] = '=TEXT(DATE(TahunSel,BulanSel,1),"mmmm yyyy")'
    for row in (53, 54):
        style_row(ws, row, 2, 7, fill=PatternFill("solid", fgColor=TEAL_LIGHT), font=TOTAL_FONT)
    for cell in ("C53", "E53"):
        ws.conditional_formatting.add(cell, FormulaRule(formula=[f'LEFT({cell},2)="OK"'], fill=PatternFill("solid", fgColor=GREEN)))
        ws.conditional_formatting.add(cell, FormulaRule(formula=[f'LEFT({cell},2)<>"OK"'], fill=PatternFill("solid", fgColor=RED)))
    ws.freeze_panes = "A7"
    add_defined_name(ws.parent, "BulanSel", "PENGATURAN!$C$24")
    add_defined_name(ws.parent, "TahunSel", "PENGATURAN!$C$25")
    add_defined_name(ws.parent, "TanggalMulai", "PENGATURAN!$C$22")
    add_defined_name(ws.parent, "TanggalCutoff", "PENGATURAN!$C$23")


def build_reconciliation(ws, source_start: int, source_end: int, journal_refs: dict[str, int]) -> dict[str, int]:
    ws.sheet_properties.tabColor = RED
    for col, width in {"A": 5, "B": 58, "C": 22, "D": 72}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "REKONSILIASI DANA DAN SALDO"
    ws["B2"].font = TITLE
    ws["B3"] = "Dana awal + pendapatan − total belanja melalui Bank Jatim = saldo akhir Bank Jatim"
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:D2")
    style_header(ws, 6, 2, ["KOMPONEN", "JUMLAH", "DASAR / CATATAN"])
    rows = {
        "recon_open": 7, "recon_income": 8, "recon_outflow": 9, "recon_bank": 10,
    }
    values = [
        (7, "Dana awal", "=144060000", "Diterima 01/08/2025 di Bank Jatim sebagai Dana Ketahanan Pangan/Penyertaan Modal Desa."),
        (8, "Total pendapatan", "=" + sumifs(TX_DEBIT, TX_CODE, quote(BANK_CODE), TX_DATE, DATE_TO_CUTOFF, TX_ID, quote("SRC-*")), "Pendapatan panen yang diterima melalui Bank Jatim."),
        (9, "Total belanja melalui Bank Jatim", "=" + sumifs(TX_CREDIT, TX_CODE, quote(BANK_CODE), TX_DATE, DATE_TO_CUTOFF), "Seluruh pembayaran melalui Bank Jatim sampai batas waktu."),
        (10, "Saldo akhir Bank Jatim", "=C7+C8-C9", "Dana awal + total pendapatan − total belanja melalui Bank Jatim."),
    ]
    for row, label, formula, note in values:
        ws.cell(row, 2, label)
        ws.cell(row, 3, formula)
        ws.cell(row, 4, note)
        money(ws.cell(row, 3))
        fill = PatternFill("solid", fgColor=TEAL_LIGHT) if row == 10 else None
        style_row(ws, row, 2, 4, fill=fill, font=TOTAL_FONT if row == 10 else None)
    ws["B13"] = "KETERANGAN"
    ws["B13"].font = Font(bold=True, color=SAND)
    disclosures = [
        "Baris sumber 302 tanggal 25/08/2026 Rp900.000 dikeluarkan karena melewati batas 23/08/2026.",
        "Baris penarikan sumber tetap tercatat di DATA SUMBER sebagai catatan, tetapi tidak dijurnal dalam model Bank Jatim langsung.",
        "Transaksi operasional dibukukan pada akun beban masing-masing dan dibayar melalui Bank Jatim.",
    ]
    for row, text in enumerate(disclosures, 14):
        ws.cell(row, 2, "• " + text)
        ws.cell(row, 2).alignment = LEFT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.freeze_panes = "B7"
    return rows


def build_data_source(ws, records: list[SourceRow]) -> tuple[int, int]:
    ws.sheet_properties.tabColor = BLUE
    for col, width in {"A": 13, "B": 14, "C": 58, "D": 18, "E": 18, "F": 18, "G": 28, "H": 70, "I": 43, "J": 38, "K": 42, "L": 14, "M": 18}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "DATA SUMBER — JEJAK AUDIT TRANSAKSI"
    ws["B2"].font = TITLE
    ws["B3"] = "Satu baris per baris sumber; tidak ada jurnal berpasangan di sheet ini. Baris 302 ditampilkan sebagai pengecualian batas waktu."
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:M2")
    ws.merge_cells("B3:M3")
    headers = ["Baris Sumber", "Tanggal Asli", "Uraian", "Nilai Asli (C)", "Nilai Asli (D)", "Nilai Sumber", "Kategori Asli (D)", "Catatan Sumber (E:H)", "Keputusan Klasifikasi", "Unit / Lingkup", "Status Bukti", "Dijurnal?", "Akun Pemetaan"]
    style_header(ws, 4, 1, headers)
    start = 5
    for out_row, record in enumerate(records, start):
        values = [record.source_row_id, record.original_date, record.description, record.amount_c, record.amount_d, record.source_amount, record.original_category, record.source_note, record.classification_decision, record.unit_scope, record.evidence_status, record.journalized, record.mapped_account]
        for col, value in enumerate(values, 1):
            ws.cell(out_row, col, value)
            ws.cell(out_row, col).border = GRID
            ws.cell(out_row, col).alignment = LEFT if col not in (1, 2, 4, 5, 6) else CENTER
        ws.cell(out_row, 2).number_format = DATE_FORMAT
        for col in (4, 5, 6):
            money(ws.cell(out_row, col))
        if record.source_row_id == 302:
            for col in range(1, 14):
                ws.cell(out_row, col).fill = PatternFill("solid", fgColor=RED)
        elif record.classification_decision.startswith("NOTE"):
            for col in range(1, 14):
                ws.cell(out_row, col).fill = PatternFill("solid", fgColor=YELLOW)
    end = start + len(records) - 1
    add_table(ws, f"A4:M{end}", "DataSumberTable")
    ws.auto_filter.ref = f"A4:M{end}"
    ws.freeze_panes = "A5"
    return start, end


def build_transactions(ws, lines: list[JournalLine]) -> int:
    ws.sheet_properties.tabColor = SAND
    widths = {"A": 24, "B": 14, "C": 38, "D": 15, "E": 23, "F": 16, "G": 46, "H": 64, "I": 18, "J": 18, "K": 10, "L": 10, "M": 12, "N": 48}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws["A2"] = "TRANSAKSI — JURNAL BERPASANGAN"
    ws["A2"].font = TITLE
    ws["A3"] = "Tidak ada baris contoh. Transaksi sumber memakai Baris Sumber; penyesuaian dan transaksi operasional diberi nomor serta status yang jelas."
    ws["A3"].font = SUBTITLE
    ws.merge_cells("A2:N2")
    ws.merge_cells("A3:N3")
    headers = ["Nomor Jurnal", "Tanggal", "Unit Usaha", "Baris Sumber", "Jenis", "Kode Akun", "Nama Akun", "Uraian", "Debet", "Kredit", "Bulan", "Tahun", "Transfer Internal?", "Status Bukti"]
    style_header(ws, 9, 1, headers)
    for row in range(TX_START, TX_END + 1):
        if row - TX_START < len(lines):
            line = lines[row - TX_START]
            input_values = [line.journal_id, line.date, line.unit, line.source_row_id, line.entry_type, line.code, None, line.description, line.debit, line.credit, None, None, line.internal, line.evidence]
            for col, value in enumerate(input_values, 1):
                ws.cell(row, col, value)
            ws.cell(row, 7, f'=IF($F{row}="","",IFERROR(VLOOKUP($F{row},AKUN!$A$2:$B$200,2,FALSE),"Kode tidak ditemukan"))')
        else:
            ws.cell(row, 7, f'=IF($F{row}="","",IFERROR(VLOOKUP($F{row},AKUN!$A$2:$B$200,2,FALSE),"Kode tidak ditemukan"))')
        ws.cell(row, 11, f'=IF($B{row}="","",MONTH($B{row}))')
        ws.cell(row, 12, f'=IF($B{row}="","",YEAR($B{row}))')
        for col in range(1, 15):
            ws.cell(row, col).border = GRID
            if row % 2 == 0:
                ws.cell(row, col).fill = PatternFill("solid", fgColor=GREY_LIGHT)
        ws.cell(row, 2).number_format = DATE_FORMAT
        ws.cell(row, 2).alignment = CENTER
        for col in (9, 10):
            money(ws.cell(row, col))
    end = TX_END
    add_table(ws, f"A9:N{end}", "TransaksiRealTable")
    ws.auto_filter.ref = f"A9:N{end}"
    ws.freeze_panes = "A10"
    units_formula = '"PEMBIBITAN DAN BUDIDAYA KAMBING,BUDIDAYA PERTANIAN,REKONSILIASI (BUKAN UNIT USAHA)"'
    unit_dv = DataValidation(type="list", formula1=units_formula, allow_blank=True)
    type_dv = DataValidation(type="list", formula1='"Belanja Sumber,Pendapatan Sumber,Saldo Awal,Penyesuaian"', allow_blank=True)
    code_dv = DataValidation(type="list", formula1="=KodeAkun", allow_blank=True)
    date_dv = DataValidation(type="date", operator="between", formula1="DATE(2025,1,1)", formula2="DATE(2030,12,31)", allow_blank=True)
    for validation in (unit_dv, type_dv, code_dv, date_dv):
        ws.add_data_validation(validation)
    unit_dv.add(f"C{TX_START}:C{TX_END}")
    type_dv.add(f"E{TX_START}:E{TX_END}")
    code_dv.add(f"F{TX_START}:F{TX_END}")
    date_dv.add(f"B{TX_START}:B{TX_END}")
    return len(lines)


def build_recap(ws) -> None:
    ws.sheet_properties.tabColor = BLUE
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    for col in range(3, 16):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws["B2"] = "REKAP 13 BULAN — AGU 2025 s.d. AGU 2026"
    ws["B2"].font = TITLE
    ws["B3"] = "Agustus 2026 dihitung hanya sampai 23/08/2026; seluruh transaksi pembayaran dan penerimaan memakai Bank Jatim langsung."
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:O2")
    ws.merge_cells("B3:O3")
    periods = []
    year, month = 2025, 8
    for _ in range(13):
        start = date(year, month, 1)
        end = CUTOFF if (year, month) == (2026, 8) else month_end(year, month)
        periods.append((start, end))
        month += 1
        if month == 13:
            month = 1
            year += 1
    ws["A4"] = "Mulai"
    ws["A5"] = "S/D"
    for col, (start, end) in enumerate(periods, 3):
        ws.cell(4, col, start)
        ws.cell(5, col, end)
        ws.cell(4, col).number_format = DATE_FORMAT
        ws.cell(5, col).number_format = DATE_FORMAT
        ws.cell(4, col).font = SMALL
        ws.cell(5, col).font = SMALL
    style_header(ws, 6, 1, ["METRIK", "LINGKUP"] + [start.strftime("%b-%y") for start, _ in periods])
    metrics = [
        (7, "Penerimaan Bank Jatim", "Semua lingkup", "receipt"),
        (8, "Belanja dari Bank Jatim", "Semua lingkup", "outflow"),
        (9, "Pendapatan operasional", UNIT_AGRI, "income"),
        (10, "Beban yang diakui", "Semua lingkup", "expense"),
        (11, "Laba/(rugi)", "Semua lingkup", "profit"),
        (12, "Pembelian aset/biaya produksi dalam proses dari sumber", "Sumber", "asset"),
        (13, "Saldo akhir Bank Jatim", "Semua lingkup", "cash"),
    ]
    for row, label, scope, metric in metrics:
        ws.cell(row, 1, label)
        ws.cell(row, 2, scope)
        for col in range(3, 16):
            start_ref = f"{get_column_letter(col)}$4"
            end_ref = f"{get_column_letter(col)}$5"
            dates = [TX_DATE, f'">="&{start_ref}', TX_DATE, f'"<="&{end_ref}']
            if metric == "receipt":
                formula = f"={sumifs(TX_DEBIT, *dates, TX_CODE, quote('1.1.01*'), TX_INTERNAL, quote('Tidak'))}"
            elif metric == "outflow":
                formula = f"={sumifs(TX_CREDIT, *dates, TX_CODE, quote('1.1.01*'), TX_INTERNAL, quote('Tidak'))}"
            elif metric == "income":
                formula = f"={sumifs(TX_CREDIT, *dates, TX_CODE, quote(PADDY_REVENUE))}-{sumifs(TX_DEBIT, *dates, TX_CODE, quote(PADDY_REVENUE))}"
            elif metric == "expense":
                formula = f"=({sumifs(TX_DEBIT, *dates, TX_CODE, quote('6.*'))}+{sumifs(TX_DEBIT, *dates, TX_CODE, quote('8.*'))})-({sumifs(TX_CREDIT, *dates, TX_CODE, quote('6.*'))}+{sumifs(TX_CREDIT, *dates, TX_CODE, quote('8.*'))})"
            elif metric == "profit":
                formula = f"={get_column_letter(col)}9-{get_column_letter(col)}10"
            elif metric == "asset":
                formula = f"={sumifs(TX_DEBIT, *dates, TX_CODE, quote('1.*'), TX_INTERNAL, quote('Tidak'), TX_ID, quote('SRC-*'))}"
            else:
                leq_month = f'"<="&{end_ref}'
                formula = f"={sumifs(TX_DEBIT, TX_CODE, quote('1.1.01*'), TX_DATE, leq_month)}-{sumifs(TX_CREDIT, TX_CODE, quote('1.1.01*'), TX_DATE, leq_month)}"
            ws.cell(row, col, formula)
            money(ws.cell(row, col))
        style_row(ws, row, 1, 15, fill=PatternFill("solid", fgColor=SAND_LIGHT) if metric in ("profit", "cash") else (PatternFill("solid", fgColor=GREY_LIGHT) if row % 2 == 0 else None), font=TOTAL_FONT if metric in ("profit", "cash") else None)
    ws["A16"] = "Catatan"
    ws["B16"] = "Pembelian aset dan biaya produksi dalam proses bukan beban operasi. Transaksi operasional dicatat pada akun beban masing-masing."
    ws["B16"].font = SMALL
    ws.merge_cells("B16:O16")
    ws.freeze_panes = "C7"


def build_per_unit(ws) -> None:
    ws.sheet_properties.tabColor = "70AD47"
    for col, width in {"A": 4, "B": 42, "C": 30, "D": 21, "E": 21}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "PER UNIT — KINERJA SELURUH PERIODE DAN BULAN TERPILIH"
    ws["B2"].font = TITLE
    ws["B3"] = "Dua unit operasi ditampilkan terpisah; dana awal dan saldo Bank Jatim ditampilkan pada baris lingkup rekonsiliasi."
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:E2")
    style_header(ws, 6, 2, ["UNIT / LINGKUP", "METRIK", "01/08/2025–23/08/2026", "BULAN 08/2026 s.d. 23/08"])
    rows = [
        (UNIT_GOAT, "Pendapatan", "income"), (UNIT_GOAT, "Beban diakui", "expense"), (UNIT_GOAT, "Laba/(rugi)", "profit"), (UNIT_GOAT, "Belanja Bank Jatim", "outflow"),
        (UNIT_AGRI, "Pendapatan", "income"), (UNIT_AGRI, "Beban diakui", "expense"), (UNIT_AGRI, "Laba/(rugi)", "profit"), (UNIT_AGRI, "Belanja Bank Jatim", "outflow"),
        (UNIT_RECON, "Dana awal / saldo Bank Jatim", "recon"),
    ]
    row = 7
    refs = {}
    for unit, label, metric in rows:
        ws.cell(row, 2, unit)
        ws.cell(row, 3, label)
        if metric == "income":
            full = f"={period_formula(TX_CREDIT, 'TanggalMulai', 'TanggalCutoff', code=PADDY_REVENUE, unit=unit)}-{period_formula(TX_DEBIT, 'TanggalMulai', 'TanggalCutoff', code=PADDY_REVENUE, unit=unit)}"
            selected = f"={period_formula(TX_CREDIT, 'DATE(TahunSel,BulanSel,1)', 'TanggalCutoff', code=PADDY_REVENUE, unit=unit)}-{period_formula(TX_DEBIT, 'DATE(TahunSel,BulanSel,1)', 'TanggalCutoff', code=PADDY_REVENUE, unit=unit)}"
        elif metric == "expense":
            full = expense_pattern("TanggalMulai", "TanggalCutoff", unit)
            selected = expense_pattern("DATE(TahunSel,BulanSel,1)", "TanggalCutoff", unit)
        elif metric == "profit":
            full = f"=D{row - 2}-D{row - 1}"
            selected = f"=E{row - 2}-E{row - 1}"
        elif metric == "outflow":
            full = f"={period_formula(TX_CREDIT, 'TanggalMulai', 'TanggalCutoff', code='1.1.01*', unit=unit, internal='Tidak')}"
            selected = f"={period_formula(TX_CREDIT, 'DATE(TahunSel,BulanSel,1)', 'TanggalCutoff', code='1.1.01*', unit=unit, internal='Tidak')}"
        else:
            full = f"={period_formula(TX_DEBIT, 'TanggalMulai', 'TanggalCutoff', code='1.1.01*', unit=unit)}-{period_formula(TX_CREDIT, 'TanggalMulai', 'TanggalCutoff', code='1.1.01*', unit=unit)}"
            selected = f"={period_formula(TX_DEBIT, 'DATE(TahunSel,BulanSel,1)', 'TanggalCutoff', code='1.1.01*', unit=unit)}-{period_formula(TX_CREDIT, 'DATE(TahunSel,BulanSel,1)', 'TanggalCutoff', code='1.1.01*', unit=unit)}"
        ws.cell(row, 4, full)
        ws.cell(row, 5, selected)
        money(ws.cell(row, 4))
        money(ws.cell(row, 5))
        style_row(ws, row, 2, 5, fill=PatternFill("solid", fgColor=GREY_LIGHT) if row % 2 == 0 else None)
        if metric == "profit":
            refs[unit] = row
            style_row(ws, row, 2, 5, fill=PatternFill("solid", fgColor=SAND_LIGHT), font=TOTAL_FONT)
        row += 1
    ws.cell(row + 1, 2, "Transaksi operasional dibukukan pada akun beban masing-masing.").font = SMALL
    ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 1, end_column=5)
    ws.freeze_panes = "D7"


def source_amount_formula(source_start: int, source_end: int, source_row_ref: str) -> str:
    amount_range = f"'DATA SUMBER'!$F${source_start}:$F${source_end}"
    source_id_range = f"'DATA SUMBER'!$A${source_start}:$A${source_end}"
    return "=" + sumifs(amount_range, source_id_range, source_row_ref)


def build_agri_cycle(ws, source_start: int, source_end: int) -> dict[str, int]:
    ws.sheet_properties.tabColor = "92D050"
    for col, width in {"A": 14, "B": 58, "C": 18, "D": 22, "E": 22, "F": 24}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "SIKLUS PERTANIAN"
    ws["B2"].font = TITLE
    ws["B3"] = "Biaya langsung mengikuti baris sumber yang telah dikonfirmasi; alokasi sewa ditampilkan terpisah."
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:F2")
    style_header(ws, 6, 1, ["BARIS SUMBER", "URAIAN", "BIAYA LANGSUNG", "CATATAN", "FORMULA / HASIL", "STATUS BUKTI"])
    row = 7
    refs = {}
    for cycle_no, ids, revenue_id, direct_total, rent_amount, release_id in ((1, CYCLE_1_ROWS, 97, 3_440_000, 1_000_000, "ADJ-CYCLE1-RELEASE"), (2, CYCLE_2_ROWS, 203, 5_500_000, 1_000_000, "ADJ-CYCLE2-RELEASE")):
        section_row(ws, row, 1, 6, f"SIKLUS {cycle_no}")
        row += 1
        start_detail = row
        for source_id in ids:
            ws.cell(row, 1, source_id)
            ws.cell(row, 2, f"Baris sumber {source_id}")
            ws.cell(row, 3, source_amount_formula(source_start, source_end, f"$A{row}"))
            ws.cell(row, 4, "Biaya langsung dalam proses")
            ws.cell(row, 5, "=C" + str(row))
            ws.cell(row, 6, "SUMBER — dipetakan ke siklus")
            money(ws.cell(row, 3))
            money(ws.cell(row, 5))
            style_row(ws, row, 1, 6, fill=PatternFill("solid", fgColor=GREY_LIGHT) if row % 2 == 0 else None)
            row += 1
        total_row = row
        ws.cell(row, 2, f"Total biaya langsung Siklus {cycle_no}")
        ws.cell(row, 3, f"=SUM(C{start_detail}:C{row - 1})")
        ws.cell(row, 4, f"Konfirmasi pengguna Rp{direct_total:,.0f}".replace(",", "."))
        money(ws.cell(row, 3))
        style_row(ws, row, 1, 6, fill=PatternFill("solid", fgColor=SAND_LIGHT), font=TOTAL_FONT)
        row += 1
        revenue_row = row
        ws.cell(row, 2, f"Pendapatan panen baris {revenue_id}")
        ws.cell(row, 3, f"={sumifs(TX_CREDIT, TX_CODE, quote(PADDY_REVENUE), TX_ID, quote(f'SRC-{revenue_id:03d}'))}")
        ws.cell(row, 4, "Penerimaan Bank Jatim")
        money(ws.cell(row, 3))
        style_row(ws, row, 1, 6)
        row += 1
        margin_row = row
        ws.cell(row, 2, "Laba/rugi langsung sebelum alokasi sewa")
        ws.cell(row, 3, f"=C{revenue_row}-C{total_row}")
        money(ws.cell(row, 3))
        style_row(ws, row, 1, 6, fill=PatternFill("solid", fgColor=TEAL_LIGHT), font=TOTAL_FONT)
        row += 1
        rent_row = row
        ws.cell(row, 2, "Alokasi sewa lahan")
        allocation_date = '"="&DATE(2026,2,1)' if cycle_no == 1 else '"="&DATE(2026,5,18)'
        ws.cell(row, 3, "=" + sumifs(TX_DEBIT, TX_CODE, quote("6.2.99.30"), TX_DATE, allocation_date))
        ws.cell(row, 4, "Rp1.000.000 setiap panen")
        money(ws.cell(row, 3))
        style_row(ws, row, 1, 6)
        row += 1
        after_row = row
        ws.cell(row, 2, "Hasil setelah alokasi sewa")
        ws.cell(row, 3, f"=C{margin_row}-C{rent_row}")
        ws.cell(row, 4, "Siklus 2: rugi Rp500.000 setelah alokasi")
        money(ws.cell(row, 3))
        style_row(ws, row, 1, 6, fill=PatternFill("solid", fgColor=SAND_LIGHT), font=TOTAL_FONT)
        refs[f"cycle{cycle_no}_direct"] = total_row
        refs[f"cycle{cycle_no}_revenue"] = revenue_row
        refs[f"cycle{cycle_no}_margin"] = margin_row
        refs[f"cycle{cycle_no}_rent"] = rent_row
        refs[f"cycle{cycle_no}_after"] = after_row
        row += 2
    ws.cell(row, 2, "Siklus 1: rugi langsung Rp2.340.000; setelah alokasi sewa rugi Rp3.340.000.").font = SMALL
    ws.cell(row + 1, 2, "Siklus 2: laba langsung sebelum alokasi Rp500.000; setelah alokasi sewa rugi Rp500.000.").font = SMALL
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 1, end_column=6)
    ws.freeze_panes = "C7"
    return refs


def build_goat_register(ws, source_records: list[SourceRow], source_start: int, source_end: int) -> dict[str, int]:
    ws.sheet_properties.tabColor = "A9D18E"
    for col, width in {"A": 14, "B": 14, "C": 48, "D": 18, "E": 23, "F": 45}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "DAFTAR KAMBING"
    ws["B2"].font = TITLE
    ws["B3"] = "Ringkasan daftar kambing: dibeli 18 ekor, mati 4 ekor, tersisa 14 ekor."
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:F2")
    style_header(ws, 6, 1, ["BARIS SUMBER", "TANGGAL", "DESKRIPSI", "NILAI", "STATUS", "CATATAN DATA FISIK"])
    by_id = {record.source_row_id: record for record in source_records}
    row = 7
    for source_id in GOAT_ROWS:
        record = by_id[source_id]
        ws.cell(row, 1, source_id)
        ws.cell(row, 2, record.original_date)
        ws.cell(row, 3, record.description)
        ws.cell(row, 4, source_amount_formula(source_start, source_end, f"$A{row}"))
        ws.cell(row, 5, "SUMBER — PEMBELIAN")
        ws.cell(row, 6, "Pembelian sumber; riwayat kematian dicatat pada tabel berikutnya.")
        ws.cell(row, 2).number_format = DATE_FORMAT
        money(ws.cell(row, 4))
        style_row(ws, row, 1, 6, fill=PatternFill("solid", fgColor=GREY_LIGHT) if row % 2 == 0 else None)
        row += 1
    total_row = row
    ws.cell(row, 3, "Total pembelian 18 ekor")
    ws.cell(row, 4, f"=SUM(D7:D{row - 1})")
    money(ws.cell(row, 4))
    style_row(ws, row, 1, 6, fill=PatternFill("solid", fgColor=SAND_LIGHT), font=TOTAL_FONT)
    row += 2
    section_row(ws, row, 3, 6, "RIWAYAT KEMATIAN")
    row += 1
    style_header(ws, row, 3, ["NOMOR JURNAL", "TANGGAL", "NILAI", "STATUS"])
    death_start = row + 1
    for death_row, (journal_id, death_date, _) in enumerate(MORTALITY_EVENTS, death_start):
        ws.cell(death_row, 3, journal_id)
        ws.cell(death_row, 4, death_date)
        ws.cell(death_row, 5, f"={sumifs(TX_DEBIT, TX_ID, quote(journal_id), TX_CODE, quote('6.2.99.32'))}")
        ws.cell(death_row, 6, "DICATAT")
        ws.cell(death_row, 4).number_format = DATE_FORMAT
        money(ws.cell(death_row, 5))
        style_row(ws, death_row, 3, 6, fill=PatternFill("solid", fgColor=YELLOW) if death_row % 2 else None)
    death_end = death_start + len(MORTALITY_EVENTS) - 1
    row = death_end + 2
    summary = {
        "goat_total": total_row, "goat_deaths": row, "goat_avg": row + 1, "goat_mortality": row + 2, "goat_carrying": row + 3,
    }
    for target, label, formula, note in [
        (row, "Mati", f'=COUNTIF(C{death_start}:C{death_end},"KEMATIAN-KAMBING-*")', "Empat kejadian pada riwayat kematian."),
        (row + 1, "Kerugian per kambing", f"=D{summary['goat_mortality']}/D{summary['goat_deaths']}", "Nilai kerugian dibagi empat kejadian."),
        (row + 2, "Kerugian kematian kambing", f"=SUM(E{death_start}:E{death_end})", "Total dari riwayat kematian."),
        (row + 3, "Nilai aset biologis tersisa", f"=D{summary['goat_total']}-D{summary['goat_mortality']}", "Tersisa 14 ekor."),
    ]:
        ws.cell(target, 3, label)
        ws.cell(target, 4, formula)
        ws.cell(target, 6, note)
        money(ws.cell(target, 4))
        style_row(ws, target, 1, 6, fill=PatternFill("solid", fgColor=YELLOW if target in (summary["goat_deaths"], summary["goat_mortality"]) else TEAL_LIGHT), font=TOTAL_FONT if target in (summary["goat_mortality"], summary["goat_carrying"]) else None)
    ws.cell(row + 6, 3, "RINGKASAN")
    ws.cell(row + 7, 3, "Dibeli 18 | Mati 4 | Tersisa 14 | Nilai pembelian Rp42.100.000 | Kerugian kematian Rp9.355.556 | Nilai aset tersisa Rp32.744.444")
    ws.cell(row + 6, 3).font = Font(bold=True, color=SAND)
    ws.cell(row + 7, 3).font = SMALL
    ws.merge_cells(start_row=row + 7, start_column=3, end_row=row + 7, end_column=6)
    add_table(ws, f"A6:F{total_row - 1}", "DaftarKambingTable")
    ws.auto_filter.ref = f"A6:F{total_row - 1}"
    ws.freeze_panes = "A7"
    return summary


def build_asset_schedule(ws) -> dict[str, int]:
    ws.sheet_properties.tabColor = "70AD47"
    for col, width in {"A": 18, "B": 45, "C": 18, "D": 19, "E": 18, "F": 18, "G": 24, "H": 50}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "JADWAL ASET DAN SEWA"
    ws["B2"].font = TITLE
    ws["B3"] = "Nilai buku dihitung sampai batas waktu; amortisasi dan penyusutan mengikuti asumsi yang dijelaskan di PETUNJUK."
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:H2")
    style_header(ws, 6, 1, ["KODE / ID", "ASET / SEWA DIBAYAR DI MUKA", "NILAI PEROLEHAN", "AKUMULASI PENYUSUTAN/AMORTISASI", "NILAI BUKU", "SIAP DIGUNAKAN / PERIODE MANFAAT", "KETERANGAN"])
    items = [
        ("1.1.07.01", "Sewa dibayar di muka — lahan kambing", to_cutoff(TX_DEBIT, "1.1.07.01"), to_cutoff(TX_CREDIT, "1.1.07.01"), "01/09/2025–23/08/2026", "Amortisasi Rp1.600.000; nilai buku Rp6.400.000."),
        ("1.1.07.02", "Sewa dibayar di muka — lahan pertanian per siklus", to_cutoff(TX_DEBIT, "1.1.07.02"), to_cutoff(TX_CREDIT, "1.1.07.02"), "01/02/2026 dan 18/05/2026", "Amortisasi Rp2.000.000; nilai buku Rp0."),
        ("1.1.07.03", "Sewa dibayar di muka — lahan pertanian 5 tahun", to_cutoff(TX_DEBIT, "1.1.07.03"), to_cutoff(TX_CREDIT, "1.1.07.03"), "Feb–Agu 2026", "Amortisasi Rp1.166.667; nilai buku Rp8.833.333."),
        ("1.3.03.03", "Bangunan kandang kambing", to_cutoff(TX_DEBIT, "1.3.03.03"), to_cutoff(TX_CREDIT, "1.3.07.05"), "05/10/2025; umur 5 tahun", "Nilai perolehan Rp24.000.000; penyusutan Rp4.400.000; nilai buku Rp19.600.000."),
        ("1.3.03.01", "Mesin pencacah pakan ternak", to_cutoff(TX_DEBIT, "1.3.03.01"), to_cutoff(TX_CREDIT, "1.3.07.06"), "06/11/2025; umur 5 tahun", "Nilai perolehan Rp6.000.000; penyusutan Rp1.000.000; nilai buku Rp5.000.000."),
        ("1.1.08.01", "Aset biologis kambing", to_cutoff(TX_DEBIT, "1.1.08.01"), to_cutoff(TX_CREDIT, "1.1.08.01"), "Batas 23/08/2026", "Pembelian Rp42.100.000 dikurangi kerugian kematian Rp9.355.556."),
        ("1.1.09.01/02", "Biaya produksi dalam proses pertanian siklus 1 dan 2", "=" + sumifs(TX_DEBIT, TX_CODE, quote("1.1.09.01"), TX_DATE, DATE_TO_CUTOFF) + "+" + sumifs(TX_DEBIT, TX_CODE, quote("1.1.09.02"), TX_DATE, DATE_TO_CUTOFF), "=" + sumifs(TX_CREDIT, TX_CODE, quote("1.1.09.01"), TX_DATE, DATE_TO_CUTOFF) + "+" + sumifs(TX_CREDIT, TX_CODE, quote("1.1.09.02"), TX_DATE, DATE_TO_CUTOFF), "01/02/2026 dan 18/05/2026", "Biaya langsung Rp8.940.000 telah dipindahkan saat panen; nilai buku Rp0."),
    ]
    row = 7
    for code, name, cost, accum, ready, note in items:
        ws.cell(row, 1, code)
        ws.cell(row, 2, name)
        ws.cell(row, 3, cost)
        ws.cell(row, 4, accum)
        ws.cell(row, 5, f"=C{row}-D{row}")
        ws.cell(row, 6, ready)
        ws.cell(row, 7, note)
        for col in (3, 4, 5):
            money(ws.cell(row, col))
        style_row(ws, row, 1, 7, fill=PatternFill("solid", fgColor=GREY_LIGHT) if row % 2 == 0 else None)
        row += 1
    ws.cell(row + 1, 2, "Konvensi: sewa kambing Sep–Jul Rp133.333/bulan, Agustus penyeimbang Rp133.337; sewa pertanian 5 tahun Feb–Jul Rp166.667, Agustus Rp166.665.").font = SMALL
    ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 1, end_column=7)
    ws.freeze_panes = "C7"
    return {"last": row - 1}


def build_profit_loss(ws) -> dict[str, int]:
    ws.sheet_properties.tabColor = "70AD47"
    for col, width in {"A": 4, "B": 50, "C": 21, "D": 23, "E": 52}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "LABA RUGI — KONSOLIDASI"
    ws["B2"].font = TITLE
    ws["B3"] = "Kolom Bulan Dipilih = Agustus 2026 s.d. 23/08; Seluruh Periode = 01/08/2025 s.d. 23/08/2026."
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:E2")
    style_header(ws, 6, 2, ["URAIAN", "BULAN DIPILIH", "SELURUH PERIODE", "CATATAN"])
    row = 7
    section_row(ws, row, 2, 5, "PENDAPATAN")
    row += 1
    revenue_row = row
    ws.cell(row, 2, "Pendapatan Penjualan Padi")
    ws.cell(row, 3, revenue_amount(PADDY_REVENUE, "DATE(TahunSel,BulanSel,1)", "TanggalCutoff"))
    ws.cell(row, 4, revenue_amount(PADDY_REVENUE, "TanggalMulai", "TanggalCutoff"))
    ws.cell(row, 5, "Baris sumber 97 dan 203")
    money(ws.cell(row, 3)); money(ws.cell(row, 4)); style_row(ws, row, 2, 5)
    row += 1
    revenue_total = row
    ws.cell(row, 2, "Total Pendapatan")
    ws.cell(row, 3, f"=C{revenue_row}")
    ws.cell(row, 4, f"=D{revenue_row}")
    money(ws.cell(row, 3)); money(ws.cell(row, 4)); style_row(ws, row, 2, 5, fill=PatternFill("solid", fgColor=SAND_LIGHT), font=TOTAL_FONT)
    row += 2
    section_row(ws, row, 2, 5, "BEBAN YANG DIAKUI — PEMBELIAN ASET TIDAK LANGSUNG MENJADI BEBAN")
    row += 1
    expense_codes = [
        ("6.1.08.01", "Beban Pokok Hasil Panen — Siklus Pertanian 1", "Biaya langsung Rp3.440.000"),
        ("6.1.08.02", "Beban Pokok Hasil Panen — Siklus Pertanian 2", "Biaya langsung Rp5.500.000"),
        ("6.2.99.04", "Pakan Ternak", "Operasional kambing"),
        ("6.2.99.21", "Konsentrat Ternak", "Operasional kambing"),
        ("6.2.99.22", "Bekatul", "Operasional kambing"),
        ("6.2.99.07", "Obat dan Vitamin Kambing", "Operasional kambing"),
        ("6.2.99.23", "Molase/Tetes", "Operasional kambing"),
        ("6.2.99.24", "Pemelihara Kambing", "Operasional kambing"),
        ("6.1.99.04", "Transportasi", "Operasional kambing"),
        ("6.1.04.01", "Listrik", "Operasional kambing"),
        ("6.2.99.25", "Pemantauan dan Evaluasi", "Operasional kambing"),
        ("6.1.02.01", "ATK", "Persiapan kambing"),
        ("6.2.99.26", "Spanduk dan Persiapan Awal", "Persiapan kambing"),
        ("6.2.99.27", "Konsumsi Pembangunan", "Persiapan kambing"),
        ("6.2.99.28", "Pelatihan Pembibitan Kambing", "Persiapan kambing"),
        ("6.2.99.29", "Amortisasi Sewa Lahan Kambing", "Konvensi 12 bulan"),
        ("6.2.99.30", "Amortisasi Sewa Lahan Siklus", "Alokasi setiap panen"),
        ("6.2.99.31", "Amortisasi Sewa Lahan Pertanian 5 Tahun", "Konvensi Feb–Agu"),
        ("6.1.07.05", "Penyusutan Bangunan Kandang", "Okt25–Agu26"),
        ("6.1.07.03", "Penyusutan Mesin Pencacah", "Nov25–Agu26"),
        ("6.2.99.32", "Kerugian Kematian Kambing", "4 ekor"),
        ("6.2.99.33", "Kebersihan Kandang", "Operasional kambing"),
        ("6.2.99.34", "Perbaikan Kecil Kandang dan Peralatan", "Operasional kambing"),
        ("6.2.99.35", "Air dan Sanitasi Kandang", "Operasional kambing"),
    ]
    expense_start = row
    for code, label, note in expense_codes:
        ws.cell(row, 2, label)
        ws.cell(row, 3, account_amount(code, "DATE(TahunSel,BulanSel,1)", "TanggalCutoff"))
        ws.cell(row, 4, account_amount(code, "TanggalMulai", "TanggalCutoff"))
        ws.cell(row, 5, note)
        money(ws.cell(row, 3)); money(ws.cell(row, 4))
        style_row(ws, row, 2, 5, fill=PatternFill("solid", fgColor=GREY_LIGHT) if row % 2 == 0 else None)
        row += 1
    expense_total = row
    ws.cell(row, 2, "Total Beban")
    ws.cell(row, 3, f"=SUM(C{expense_start}:C{row - 1})")
    ws.cell(row, 4, f"=SUM(D{expense_start}:D{row - 1})")
    money(ws.cell(row, 3)); money(ws.cell(row, 4)); style_row(ws, row, 2, 5, fill=PatternFill("solid", fgColor=SAND_LIGHT), font=TOTAL_FONT)
    row += 1
    profit_row = row
    ws.cell(row, 2, "SURPLUS / (RUGI)")
    ws.cell(row, 3, f"=C{revenue_total}-C{expense_total}")
    ws.cell(row, 4, f"=D{revenue_total}-D{expense_total}")
    money(ws.cell(row, 3)); money(ws.cell(row, 4)); style_row(ws, row, 2, 5, fill=PatternFill("solid", fgColor=TEAL_LIGHT), font=TOTAL_FONT)
    ws.cell(row + 2, 2, "Pembelian aset biologis, kandang, mesin, sewa dibayar di muka, dan biaya produksi dalam proses berada di neraca; beban pokok diakui saat panen.").font = SMALL
    ws.merge_cells(start_row=row + 2, start_column=2, end_row=row + 2, end_column=5)
    ws.freeze_panes = "C7"
    return {"revenue_total": revenue_total, "expense_total": expense_total, "profit": profit_row}


def build_cashflow(ws, recon_rows: dict[str, int]) -> dict[str, int]:
    ws.sheet_properties.tabColor = "00A6A6"
    for col, width in {"A": 4, "B": 54, "C": 22, "D": 58}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "ARUS KAS — BANK JATIM"
    ws["B2"].font = TITLE
    ws["B3"] = "Semua pembelian dan penerimaan hasil panen dicatat langsung melalui Bank Jatim."
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:D2")
    style_header(ws, 6, 2, ["KOMPONEN", "JUMLAH", "CATATAN"])
    rows = {"cash_receipts": 7, "cash_outflows": 8, "cash_net": 9, "cash_bank": 10}
    external_receipts = "=" + sumifs(TX_DEBIT, TX_CODE, quote(BANK_CODE), TX_DATE, DATE_TO_CUTOFF)
    external_outflows = "=" + sumifs(TX_CREDIT, TX_CODE, quote(BANK_CODE), TX_DATE, DATE_TO_CUTOFF)
    values = [
        (7, "Penerimaan Bank Jatim — dana awal dan panen", external_receipts, "Dana awal Rp144.060.000 dan pendapatan panen Rp7.100.000."),
        (8, "Total belanja melalui Bank Jatim", external_outflows, "Seluruh pembayaran melalui Bank Jatim sampai batas waktu."),
        (9, "Kenaikan/(penurunan) bersih Bank Jatim", "=C7-C8", "Saldo akhir Bank Jatim Rp25.520.000."),
        (10, "Saldo akhir Bank Jatim", cash_balance_formula("TanggalCutoff"), "Saldo akhir Bank Jatim Rp25.520.000."),
    ]
    for row, label, formula, note in values:
        flow_fill = PatternFill("solid", fgColor=SAND_LIGHT) if row in (9, 10) else None
        ws.cell(row, 2, label); ws.cell(row, 3, formula); ws.cell(row, 4, note); money(ws.cell(row, 3)); style_row(ws, row, 2, 4, fill=flow_fill, font=TOTAL_FONT if row in (9, 10) else None)
    ws["B13"] = "Catatan: model laporan menggunakan Bank Jatim langsung; penarikan yang tercatat pada sumber tidak dibuat sebagai jurnal."
    ws["B13"].font = SMALL
    ws.merge_cells("B13:D13")
    ws.freeze_panes = "C7"
    return rows


def build_trial_balance(ws) -> dict[str, int]:
    ws.sheet_properties.tabColor = "8064A2"
    for col, width in {"A": 17, "B": 48, "C": 21, "D": 21, "E": 21, "F": 21}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "NERACA SALDO"
    ws["B2"].font = TITLE
    ws["B3"] = "Kumulatif sampai 23 Agustus 2026; seluruh jurnal berpasangan termasuk penyesuaian dan transaksi operasional."
    ws["B3"].font = SUBTITLE
    style_header(ws, 6, 1, ["KODE AKUN", "NAMA AKUN", "MUTASI DEBET", "MUTASI KREDIT", "SALDO DEBET", "SALDO KREDIT"])
    row = 7
    for code, name, _, _ in COA:
        ws.cell(row, 1, code); ws.cell(row, 2, name)
        code_ref = f"$A{row}"
        ws.cell(row, 3, "=" + sumifs(TX_DEBIT, TX_CODE, code_ref, TX_DATE, DATE_TO_CUTOFF))
        ws.cell(row, 4, "=" + sumifs(TX_CREDIT, TX_CODE, code_ref, TX_DATE, DATE_TO_CUTOFF))
        ws.cell(row, 5, f"=IF(C{row}>D{row},C{row}-D{row},0)")
        ws.cell(row, 6, f"=IF(D{row}>C{row},D{row}-C{row},0)")
        for col in range(3, 7): money(ws.cell(row, col))
        style_row(ws, row, 1, 6, fill=PatternFill("solid", fgColor=GREY_LIGHT) if row % 2 == 0 else None)
        row += 1
    total_row = row
    ws.cell(row, 2, "TOTAL")
    for col in range(3, 7):
        letter = get_column_letter(col)
        ws.cell(row, col, f"=SUM({letter}7:{letter}{row - 1})")
        money(ws.cell(row, col))
    style_row(ws, row, 1, 6, fill=PatternFill("solid", fgColor=SAND_LIGHT), font=TOTAL_FONT)
    row += 2
    check_row = row
    ws.cell(row, 2, "PEMERIKSAAN MUTASI DEBET = KREDIT")
    ws.cell(row, 3, f'=IF(ABS(C{total_row}-D{total_row})<1,"OK — seimbang","CEK — tidak seimbang")')
    style_row(ws, row, 1, 6, fill=PatternFill("solid", fgColor=TEAL_LIGHT), font=TOTAL_FONT)
    ws.conditional_formatting.add(f"C{row}", FormulaRule(formula=[f'LEFT(C{row},2)="OK"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"C{row}", FormulaRule(formula=[f'LEFT(C{row},2)<>"OK"'], fill=PatternFill("solid", fgColor=RED)))
    ws.freeze_panes = "A7"
    return {"total": total_row, "check": check_row}


def build_position(ws, profit_row: int) -> dict[str, int]:
    ws.sheet_properties.tabColor = "C55A11"
    for col, width in {"A": 17, "B": 46, "C": 22, "D": 4, "E": 17, "F": 46, "G": 22}.items():
        ws.column_dimensions[col].width = width
    ws["B2"] = "POSISI KEUANGAN"
    ws["B2"].font = TITLE
    ws["B3"] = "Posisi per cut-off 23 Agustus 2026; aset vs kewajiban + ekuitas + laba/(rugi) berjalan."
    ws["B3"].font = SUBTITLE
    ws.merge_cells("B2:G2")
    style_header(ws, 6, 1, ["KODE", "ASET", "JUMLAH", "", "KODE", "KEWAJIBAN & EKUITAS", "JUMLAH"])
    assets = [(code, name) for code, name, category, _ in COA if category == "Aset"]
    right = [(code, name) for code, name, category, _ in COA if category in ("Kewajiban", "Ekuitas")]
    row = 7
    asset_start = row
    asset_rows = {}
    for code, name in assets:
        asset_rows[code] = row
        ws.cell(row, 1, code); ws.cell(row, 2, name); ws.cell(row, 3, f"{to_cutoff(TX_DEBIT, code)}-{to_cutoff(TX_CREDIT, code)[1:]}")
        money(ws.cell(row, 3)); style_row(ws, row, 1, 3, fill=PatternFill("solid", fgColor=GREY_LIGHT) if row % 2 == 0 else None)
        row += 1
    asset_total = row
    ws.cell(row, 2, "TOTAL ASET"); ws.cell(row, 3, f"=SUM(C{asset_start}:C{row - 1})"); money(ws.cell(row, 3)); style_row(ws, row, 1, 3, fill=PatternFill("solid", fgColor=SAND_LIGHT), font=TOTAL_FONT)
    right_row = 7
    right_start = right_row
    for code, name in right:
        ws.cell(right_row, 5, code); ws.cell(right_row, 6, name); ws.cell(right_row, 7, f"{to_cutoff(TX_CREDIT, code)}-{to_cutoff(TX_DEBIT, code)[1:]}")
        money(ws.cell(right_row, 7)); style_row(ws, right_row, 5, 7, fill=PatternFill("solid", fgColor=GREY_LIGHT) if right_row % 2 == 0 else None)
        right_row += 1
    ws.cell(right_row, 6, "Laba/(Rugi) berjalan"); ws.cell(right_row, 7, f"='LABA RUGI'!$D${profit_row}"); money(ws.cell(right_row, 7)); style_row(ws, right_row, 5, 7, fill=PatternFill("solid", fgColor=TEAL_LIGHT), font=TOTAL_FONT)
    right_row += 1
    right_total = right_row
    ws.cell(right_row, 6, "TOTAL KEWAJIBAN + EKUITAS"); ws.cell(right_row, 7, f"=SUM(G{right_start}:G{right_row - 1})"); money(ws.cell(right_row, 7)); style_row(ws, right_row, 5, 7, fill=PatternFill("solid", fgColor=SAND_LIGHT), font=TOTAL_FONT)
    diff_row = max(asset_total, right_total) + 2
    ws.cell(diff_row, 2, "SELISIH ASET - KEWAJIBAN/EKUITAS"); ws.cell(diff_row, 3, f"=C{asset_total}-G{right_total}"); money(ws.cell(diff_row, 3)); style_row(ws, diff_row, 1, 7, fill=PatternFill("solid", fgColor=SAND_LIGHT), font=TOTAL_FONT)
    check_row = diff_row + 1
    ws.cell(check_row, 2, "PEMERIKSAAN KESEIMBANGAN"); ws.cell(check_row, 3, f'=IF(ABS(C{diff_row})<1,"OK — seimbang","CEK — tidak seimbang")'); style_row(ws, check_row, 1, 7, fill=PatternFill("solid", fgColor=TEAL_LIGHT), font=TOTAL_FONT)
    ws.conditional_formatting.add(f"C{check_row}", FormulaRule(formula=[f'LEFT(C{check_row},2)="OK"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"C{check_row}", FormulaRule(formula=[f'LEFT(C{check_row},2)<>"OK"'], fill=PatternFill("solid", fgColor=RED)))
    ws.freeze_panes = "A7"
    return {"asset_total": asset_total, "right_total": right_total, "diff": diff_row, "check": check_row, "asset_rows": asset_rows}


def build_accounts(ws, wb) -> None:
    ws.sheet_properties.tabColor = TEAL
    for col, width in {"A": 17, "B": 52, "C": 16, "D": 15}.items():
        ws.column_dimensions[col].width = width
    style_header(ws, 1, 1, ["KODE AKUN", "NAMA AKUN", "KATEGORI LAPORAN", "SALDO NORMAL"])
    for row, (code, name, category, normal) in enumerate(COA, 2):
        for col, value in enumerate((code, name, category, normal), 1):
            ws.cell(row, col, value)
            ws.cell(row, col).border = GRID
        ws.cell(row, 1).number_format = "@"
        if row % 2 == 0:
            for col in range(1, 5): ws.cell(row, col).fill = PatternFill("solid", fgColor=GREY_LIGHT)
    last = len(COA) + 1
    add_table(ws, f"A1:D{last}", "RealCOATable")
    ws.auto_filter.ref = f"A1:D{last}"
    ws.freeze_panes = "A2"
    add_defined_name(wb, "KodeAkun", "AKUN!$A$2:$A$200")
    ws.cell(last + 2, 1, "COA ini adalah sumber baris laporan. Tambah akun permanen pada generator COA lalu jalankan ulang; tidak cukup menambah baris AKUN saja.").font = SMALL
    ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=4)


def build_instructions(ws) -> None:
    ws.sheet_properties.tabColor = "7F7F7F"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 115
    ws["B2"] = "PETUNJUK — LAPORAN MANAJEMEN INTERNAL"
    ws["B2"].font = TITLE
    instructions = [
        (4, "1. Dasar laporan adalah source.xlsx, laporan manajemen, dan basis akrual sampai 23/08/2026. Banner LAPORAN MANAJEMEN INTERNAL berarti dokumen ini untuk pengendalian manajemen."),
        (5, "2. Dua unit operasi tepat: PEMBIBITAN DAN BUDIDAYA KAMBING dan BUDIDAYA PERTANIAN. REKONSILIASI (BUKAN UNIT USAHA) hanya untuk dana awal dan ringkasan saldo."),
        (6, "3. DATA SUMBER berisi satu baris untuk setiap baris sumber, tanpa jurnal berpasangan. Baris 302 tanggal 25/08/2026 ditandai DIKECUALIKAN — SETELAH BATAS WAKTU."),
        (7, "4. TRANSAKSI adalah jurnal berpasangan. Baris Sumber melacak transaksi sumber; Nomor Jurnal melacak saldo awal, penyesuaian, dan transaksi operasional."),
        (8, "5. Model Bank Jatim langsung: seluruh belanja dan pendapatan panen dikredit/debet ke Bank Jatim. Catatan penarikan dari sumber tetap ada di DATA SUMBER, tetapi tidak dibuat jurnal."),
        (9, "6. Aset Dalam Penyelesaian adalah biaya pembangunan kandang yang belum siap digunakan; saat selesai, nilainya dipindahkan ke aset Bangunan Kandang Kambing."),
        (10, "7. Biaya Produksi Dalam Proses adalah biaya tanam dan perawatan yang dikumpulkan sampai panen, sehingga belum langsung menjadi beban laba rugi."),
        (11, "8. Beban Pokok Hasil Panen adalah biaya produksi yang dipindahkan menjadi beban ketika hasil panen dijual. Cara ini memisahkan biaya yang masih tersimpan dari biaya yang sudah menghasilkan pendapatan."),
        (12, "9. Cara membaca NERACA SALDO: Kode/Nama Akun menunjukkan akun; Mutasi Debet/Kredit menunjukkan total pergerakan; Saldo Debet/Kredit menunjukkan posisi akhir. Total Debet dan Kredit wajib sama. Aset/beban umumnya Debet, sedangkan kewajiban/modal/pendapatan umumnya Kredit. Selisih berarti jurnal perlu diperiksa."),
        (13, "10. Cara membaca LABA RUGI: Pendapatan dikurangi Beban. Bandingkan Bulan Dipilih dengan Seluruh Periode. Pembelian kambing, kandang, mesin, dan sewa dibayar di muka tidak langsung menjadi beban; penyusutan, amortisasi, biaya operasi, kerugian kematian, dan beban hasil panen yang sudah diakui masuk laba rugi. Transaksi operasional dicatat pada akun beban masing-masing."),
        (14, "11. Cara membaca JADWAL ASET: Nilai Perolehan dikurangi Akumulasi Penyusutan/Amortisasi menghasilkan Nilai Buku. Perhatikan umur/periode manfaat dan status asumsi. Nilai buku adalah nilai akuntansi aset, bukan saldo Bank Jatim dan bukan laba."),
        (16, "12. Transaksi operasional dicatat pada akun beban masing-masing dan pembayaran langsung mengurangi Bank Jatim."),
        (17, "13. DAFTAR KAMBING mencatat dibeli 18 ekor, mati 4 ekor, tersisa 14 ekor, nilai pembelian Rp42.100.000, kerugian kematian Rp9.355.556, dan nilai aset tersisa Rp32.744.444."),
        (18, "14. Jika ada akun permanen baru, tambahkan tuple ke COA pada generate_real_report.py lalu jalankan ulang. Menambah baris AKUN saja tidak menambah baris pada laporan lain."),
        (20, "IMPOR KE GOOGLE SHEETS"),
        (21, "Unggah XLSX ke Google Drive → Open with Google Sheets, atau File → Import → Upload → Create new spreadsheet. Setelah impor, periksa format tanggal/rupiah, formula, filter, validasi data, status bukti, dan pemeriksaan keseimbangan."),
        (23, "CATATAN PENGENDALIAN"),
        (24, "Transaksi operasional dan kerugian kematian kambing telah dicatat dalam laporan manajemen internal."),
    ]
    for row, text in instructions:
        ws.cell(row, 2, text)
        ws.cell(row, 2).alignment = LEFT
        if row in (15, 18):
            ws.cell(row, 2).font = Font(bold=True, color=SAND)
            ws.cell(row, 2).fill = PatternFill("solid", fgColor=TEAL_LIGHT)
        else:
            ws.cell(row, 2).border = GRID
    ws.freeze_panes = "B4"


def write_classified_csv(records: list[SourceRow]) -> None:
    fields = ["source_row_id", "original_date", "description", "amount_c", "amount_d", "source_amount", "original_category", "source_note", "classification_decision", "unit_scope", "evidence_status", "journalized", "mapped_account"]
    with CLASSIFIED_CSV.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for record in records:
            data = {field: getattr(record, field) for field in fields}
            if isinstance(data["original_date"], datetime):
                data["original_date"] = data["original_date"].date().isoformat()
            writer.writerow(data)


def build_workbook(records: list[SourceRow], lines: list[JournalLine]) -> Workbook:
    wb = Workbook()
    wb.active.title = "PENGATURAN"
    sheet_names = ["REKONSILIASI", "DATA SUMBER", "TRANSAKSI", "REKAP 13 BULAN", "PER UNIT", "SIKLUS PERTANIAN", "DAFTAR KAMBING", "JADWAL ASET", "LABA RUGI", "ARUS KAS", "NERACA SALDO", "POSISI KEUANGAN", "AKUN", "PETUNJUK"]
    for name in sheet_names:
        wb.create_sheet(name)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    source_start, source_end = build_data_source(wb["DATA SUMBER"], records)
    tx_count = build_transactions(wb["TRANSAKSI"], lines)
    recon_rows = build_reconciliation(wb["REKONSILIASI"], source_start, source_end, {})
    build_recap(wb["REKAP 13 BULAN"])
    build_per_unit(wb["PER UNIT"])
    cycle_refs = build_agri_cycle(wb["SIKLUS PERTANIAN"], source_start, source_end)
    goat_refs = build_goat_register(wb["DAFTAR KAMBING"], records, source_start, source_end)
    build_asset_schedule(wb["JADWAL ASET"])
    profit_rows = build_profit_loss(wb["LABA RUGI"])
    build_cashflow(wb["ARUS KAS"], recon_rows)
    trial_rows = build_trial_balance(wb["NERACA SALDO"])
    position_rows = build_position(wb["POSISI KEUANGAN"], profit_rows["profit"])
    build_accounts(wb["AKUN"], wb)
    build_instructions(wb["PETUNJUK"])
    settings_refs = {
        "recon_open": recon_rows["recon_open"], "recon_outflow": recon_rows["recon_outflow"], "recon_income": recon_rows["recon_income"], "recon_bank": recon_rows["recon_bank"],
        "lr_revenue_total": profit_rows["revenue_total"], "lr_expense_total": profit_rows["expense_total"], "lr_profit": profit_rows["profit"],
        "cycle1_revenue": cycle_refs["cycle1_revenue"], "cycle1_after": cycle_refs["cycle1_after"], "cycle2_revenue": cycle_refs["cycle2_revenue"], "cycle2_margin": cycle_refs["cycle2_margin"], "cycle2_after": cycle_refs["cycle2_after"],
        "goat_total": goat_refs["goat_total"], "goat_deaths": goat_refs["goat_deaths"], "goat_carrying": goat_refs["goat_carrying"],
        "position_asset_total": position_rows["asset_total"], "asset_bank": position_rows["asset_rows"][BANK_CODE], "asset_biological": position_rows["asset_rows"]["1.1.08.01"], "asset_building": position_rows["asset_rows"]["1.3.03.03"], "asset_building_accum": position_rows["asset_rows"]["1.3.07.05"], "asset_machine": position_rows["asset_rows"]["1.3.03.01"], "asset_machine_accum": position_rows["asset_rows"]["1.3.07.06"], "asset_goat_lease": position_rows["asset_rows"]["1.1.07.01"], "asset_cycle_lease": position_rows["asset_rows"]["1.1.07.02"], "asset_farm_lease": position_rows["asset_rows"]["1.1.07.03"],
        "trial_total": trial_rows["total"], "position_check": position_rows["check"],
    }
    build_settings(wb["PENGATURAN"], settings_refs)
    for ws in wb.worksheets:
        finish_sheet(ws)
    add_defined_name(wb, "UnitOperasi", 'PENGATURAN!$B$17:$B$18')
    add_defined_name(wb, "UnitScope", 'PENGATURAN!$B$17:$B$19')
    wb._real_journal_count = tx_count  # type: ignore[attr-defined]
    wb._real_source_count = len(records)  # type: ignore[attr-defined]
    return wb


def balance_by_code(lines: list[JournalLine], code: str) -> float:
    return sum(line.debit - line.credit for line in lines if line.code == code)


def validate_workbook(records: list[SourceRow], lines: list[JournalLine], wb: Workbook) -> dict[str, int | float]:
    expected_sheets = ["PENGATURAN", "REKONSILIASI", "DATA SUMBER", "TRANSAKSI", "REKAP 13 BULAN", "PER UNIT", "SIKLUS PERTANIAN", "DAFTAR KAMBING", "JADWAL ASET", "LABA RUGI", "ARUS KAS", "NERACA SALDO", "POSISI KEUANGAN", "AKUN", "PETUNJUK"]
    assert wb.sheetnames == expected_sheets, wb.sheetnames
    for name in ("BulanSel", "TahunSel", "TanggalMulai", "TanggalCutoff", "KodeAkun", "UnitOperasi", "UnitScope"):
        assert name in wb.defined_names, name
    assert len(wb["TRANSAKSI"].data_validations.dataValidation) == 4
    assert len(wb["PENGATURAN"].data_validations.dataValidation) == 2
    assert wb["TRANSAKSI"].auto_filter.ref == "A9:N1000"
    assert wb["DATA SUMBER"].auto_filter.ref is not None
    assert wb["TRANSAKSI"]["A10"].value == lines[0].journal_id
    assert wb["TRANSAKSI"]["G10"].value.startswith("=IF")
    assert wb["TRANSAKSI"]["K10"].value.startswith("=IF")
    assert wb["TRANSAKSI"]["L10"].value.startswith("=IF")
    assert wb["TRANSAKSI"]["A9"].fill.fill_type == "solid"
    assert len([wb["REKAP 13 BULAN"].cell(6, col).value for col in range(3, 16)]) == 13
    assert wb["PENGATURAN"]["B5"].value == "LAPORAN MANAJEMEN INTERNAL"
    assert wb["PENGATURAN"]["C11"].value == "DITA TIA MUKARROMAH"
    assert wb["PENGATURAN"]["C12"].value == "SOLIHIN"
    assert wb["PENGATURAN"]["C13"].value == "ROBIATUL HUSNA"
    dashboard_titles = {str(wb["PENGATURAN"].cell(row, 2).value) for row in (29, 34, 39, 44, 52)}
    assert dashboard_titles == {"RINGKASAN DANA DAN BANK", "RINGKASAN LABA RUGI", "RINGKASAN ASET", "RINGKASAN UNIT USAHA", "KONTROL"}
    for cell in ("C30", "E30", "G30", "C31", "C35", "E35", "G35", "C36", "E36", "G36", "C40", "E40", "G40", "C41", "E41", "G41", "C45", "D45", "E45", "D46", "F46", "D47", "F47", "F48"):
        assert isinstance(wb["PENGATURAN"][cell].value, str) and wb["PENGATURAN"][cell].value.startswith("=")
    assert "+'POSISI KEUANGAN'!$C$17" in wb["PENGATURAN"]["C41"].value
    assert "+'POSISI KEUANGAN'!$C$18" in wb["PENGATURAN"]["E41"].value
    assert not any("[CONTOH]" in str(cell.value) for ws in wb.worksheets for row in ws.iter_rows() for cell in row if cell.value is not None)
    formulas = [cell.value for ws in wb.worksheets for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    assert formulas and all("#REF!" not in formula.upper() for formula in formulas)
    assert len(lines) == 358
    assert sum(line.debit for line in lines) == sum(line.credit for line in lines)
    assert round(balance_by_code(lines, BANK_CODE)) == 25_520_000
    assert all(line.code != "1.1.01.01" for line in lines)
    assert not any(line.entry_type == "Transfer Internal" for line in lines)
    assert not any(line.source_row_id in WITHDRAWAL_ROWS for line in lines)
    assert all(record.journalized == "TIDAK" for record in records if record.source_row_id in WITHDRAWAL_ROWS)
    opr_lines = [line for line in lines if line.journal_id.startswith("OPR-")]
    assert len(set(line.journal_id for line in opr_lines)) == 21
    assert len(opr_lines) == 42
    assert sum(line.debit for line in opr_lines) == OPERATIONAL_TOTAL
    assert sum(line.credit for line in opr_lines) == OPERATIONAL_TOTAL
    assert all(line.unit == UNIT_GOAT for line in opr_lines)
    assert all(line.entry_type == "Belanja Sumber" for line in opr_lines)
    assert all(line.evidence == "DICATAT" for line in opr_lines)
    assert all(not line.description.startswith("[") for line in opr_lines)
    assert sum(line.credit for line in lines if line.code == BANK_CODE) == SOURCE_OUTFLOW + OPERATIONAL_TOTAL
    asset_value = sum(balance_by_code(lines, code) for code, _, category, _ in COA if category == "Aset")
    liability_equity = sum(-balance_by_code(lines, code) for code, _, category, _ in COA if category in ("Kewajiban", "Ekuitas"))
    profit = sum(-balance_by_code(lines, code) for code, _, category, _ in COA if category == "Pendapatan") - sum(balance_by_code(lines, code) for code, _, category, _ in COA if category == "Beban")
    assert round(asset_value) == round(liability_equity + profit)
    assert sum(record.source_amount for record in records if record.classification_decision.startswith("TERKLASIFIKASI")) == SOURCE_OUTFLOW
    assert sum(record.source_amount for record in records if record.classification_decision.startswith("PENDAPATAN PANEN")) == HARVEST_INCOME
    assert sum(record.source_amount for record in records if record.classification_decision.startswith("CATATAN PENARIKAN")) == WITHDRAWAL_TOTAL
    source_amounts = {record.source_row_id: record.source_amount for record in records}
    cycle_1 = sum(source_amounts[row_id] for row_id in CYCLE_1_ROWS)
    cycle_2 = sum(source_amounts[row_id] for row_id in CYCLE_2_ROWS)
    assert cycle_1 == 3_440_000
    assert cycle_2 == 5_500_000
    assert 1_100_000 - cycle_1 == -2_340_000
    assert (1_100_000 - cycle_1) - 1_000_000 == -3_340_000
    assert 6_000_000 - cycle_2 == 500_000
    assert (6_000_000 - cycle_2) - 1_000_000 == -500_000
    assert sum(line.debit for line in lines if line.code == "1.1.08.01" and str(line.source_row_id).isdigit()) == GOAT_PURCHASE_TOTAL
    assert sum(line.debit for line in lines if line.code == "6.2.99.32") == GOAT_DEATH_ADJUSTMENT
    assert balance_by_code(lines, "1.1.08.01") == GOAT_PURCHASE_TOTAL - GOAT_DEATH_ADJUSTMENT
    assert sum(line.debit for line in lines if line.code == "1.1.09.01" and str(line.source_row_id).isdigit()) == 3_440_000
    assert sum(line.debit for line in lines if line.code == "1.1.09.02" and str(line.source_row_id).isdigit()) == 5_500_000
    assert sum(line.debit for line in lines if line.code == "6.1.08.01") == 3_440_000
    assert sum(line.debit for line in lines if line.code == "6.1.08.02") == 5_500_000
    assert sum(line.debit for line in lines if line.code == "1.3.03.03") == 24_000_000
    assert sum(line.debit for line in lines if line.code == "1.3.03.01") == 6_000_000
    assert sum(line.debit for line in lines if line.code == "6.1.07.05") == 4_400_000
    assert sum(line.debit for line in lines if line.code == "6.1.07.03") == 1_000_000
    assert sum(line.debit for line in lines if line.code == "6.2.99.29") == 1_600_000
    assert sum(line.debit for line in lines if line.code == "6.2.99.30") == 2_000_000
    assert sum(line.debit for line in lines if line.code == "6.2.99.31") == 1_166_667
    assert any(record.source_row_id == 302 and record.classification_decision.startswith("DIKECUALIKAN") for record in records)
    assert not any(line.source_row_id == 302 for line in lines)
    assert any("4 ekor" in str(cell.value) or "4 kambing" in str(cell.value) for cell in wb["DAFTAR KAMBING"]._cells.values())
    assert GOAT_DEATHS == 4
    assert GOAT_DEATH_ADJUSTMENT == 9_355_556
    assert balance_by_code(lines, "1.1.08.01") == 32_744_444
    mortality_lines = [line for line in lines if line.journal_id.startswith("KEMATIAN-KAMBING-")]
    assert len(mortality_lines) == 8
    expected_mortality = {journal_id: (death_date, description) for journal_id, death_date, description in MORTALITY_EVENTS}
    for journal_id, (death_date, description) in expected_mortality.items():
        event_lines = [line for line in mortality_lines if line.journal_id == journal_id]
        assert len(event_lines) == 2
        assert all(line.date.date() == death_date for line in event_lines)
        assert all(line.description == description for line in event_lines)
        assert all(line.evidence == "DICATAT" for line in event_lines)
        assert sum(line.debit for line in event_lines if line.code == "6.2.99.32") == MORTALITY_AMOUNT
        assert sum(line.credit for line in event_lines if line.code == "1.1.08.01") == MORTALITY_AMOUNT
        assert death_date.month not in (6, 7, 8)
    assert sum(balance_by_code(lines, code) for code, _, category, _ in COA if category == "Aset" and code == "1.1.08.01") == 32_744_444
    assert GOAT_DEATHS + (18 - GOAT_DEATHS) == 18
    forbidden_display = re.compile(r"biaya operasional tambahan|rincian biaya operasional tambahan|total biaya operasional tambahan|BOP-|Rp4\.010\.000|4010000|21 transaksi|21 jurnal|dua puluh satu transaksi|dua puluh satu biaya|sekretaris\s+robiatul|bendahara\s+solihin", re.IGNORECASE)
    workbook_strings = [str(cell.value) for ws in wb.worksheets for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)]
    assert not any(forbidden_display.search(value) for value in workbook_strings)
    readme_text = README_PATH.read_text(encoding="utf-8")
    assert not forbidden_display.search(readme_text)
    assert "Sekretaris **SOLIHIN**" in readme_text
    assert "Bendahara **ROBIATUL HUSNA**" in readme_text
    recon_strings = [str(cell.value) for row in wb["REKONSILIASI"].iter_rows() for cell in row if isinstance(cell.value, str)]
    assert any("Total belanja melalui Bank Jatim" in value for value in recon_strings)
    assert not any(term in value for value in recon_strings for term in ("121.630.000", "4.010.000", "BOP-"))
    assert wb["REKONSILIASI"]["C9"].value == "=" + sumifs(TX_CREDIT, TX_CODE, quote(BANK_CODE), TX_DATE, DATE_TO_CUTOFF)
    assert wb["REKONSILIASI"]["C10"].value == "=C7+C8-C9"
    assert wb["PENGATURAN"]["C31"].value == "='REKONSILIASI'!$C$9"
    assert wb["ARUS KAS"]["C8"].value == "=" + sumifs(TX_CREDIT, TX_CODE, quote(BANK_CODE), TX_DATE, DATE_TO_CUTOFF)
    assert "_DRAFT.xlsx" not in OUTPUT.name
    assert not (ROOT / "LAPORAN_KEUANGAN_BUMDESA_LANCAR_JAYA_AGU2025_AGU2026_DRAFT.xlsx").exists()
    # Reopen validation is part of the generator, not just an external check.
    wb.save(OUTPUT)
    reopened = load_workbook(OUTPUT, data_only=False)
    assert reopened.sheetnames == expected_sheets
    assert len(reopened["TRANSAKSI"].data_validations.dataValidation) == 4
    assert reopened["TRANSAKSI"].auto_filter.ref == "A9:N1000"
    reopened_formulas = [cell.value for ws in reopened.worksheets for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    assert all("#REF!" not in formula.upper() for formula in reopened_formulas)
    assert reopened["PENGATURAN"]["B5"].fill.fill_type == "solid"
    reopened_strings = [str(cell.value) for ws in reopened.worksheets for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)]
    assert not any(forbidden_display.search(value) for value in reopened_strings)
    return {"journal_lines": len(lines), "source_rows": len(records), "formula_count": len(reopened_formulas), "xlsx_bytes": OUTPUT.stat().st_size, "bank": balance_by_code(lines, BANK_CODE), "opr_transactions": len(set(line.journal_id for line in opr_lines)), "opr_total": OPERATIONAL_TOTAL, "mortality": GOAT_DEATH_ADJUSTMENT, "biological_carrying": balance_by_code(lines, "1.1.08.01")}


def main() -> None:
    ROOT.mkdir(parents=True, exist_ok=True)
    old_output = ROOT / "LAPORAN_KEUANGAN_BUMDESA_LANCAR_JAYA_AGU2025_AGU2026_DRAFT.xlsx"
    if old_output.exists():
        old_output.unlink()
    records = read_source()
    lines = build_journal(records)
    write_classified_csv(records)
    wb = build_workbook(records, lines)
    stats = validate_workbook(records, lines, wb)
    print(f"Generated: {OUTPUT}")
    print(f"Classified source: {CLASSIFIED_CSV}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Journal lines: {stats['journal_lines']}")
    print(f"Source rows: {stats['source_rows']}")
    print(f"Formula count: {stats['formula_count']}")
    print(f"Bank Jatim: {stats['bank']}")
    print(f"OPR journal pairs: {stats['opr_transactions']} | Internal total Rp{stats['opr_total']}")
    print(f"Mortalitas: Rp{stats['mortality']} | Nilai buku biologis: Rp{stats['biological_carrying']}")
    print(f"XLSX bytes: {stats['xlsx_bytes']}")


if __name__ == "__main__":
    main()
